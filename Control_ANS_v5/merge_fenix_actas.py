"""
------------------------------------------------------------
🔄 MERGE_FENIX_ACTAS.PY – Cruce Programación vs Actas (Versión Final)
------------------------------------------------------------
Autor: Héctor A. Gaviria + IA (2025)
------------------------------------------------------------
Descripción:
1️⃣ Cruza Programación (pendientes) vs Actas de Clientes.
2️⃣ Actualiza columna ESTADO_FENIX directamente en FENIX_ANS.xlsx.
3️⃣ Mueve pedidos cerrados (Ejecutado en Campo + Cumplido)
    al archivo REPOSITORIO_PEDIDOS_CERRADOS.xlsx.
4️⃣ Aplica formato de color en ESTADO_FENIX según días restantes.
------------------------------------------------------------
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill


# ------------------------------------------------------------
# 📂 RUTAS DE ARCHIVOS
# ------------------------------------------------------------
base_dir = Path(__file__).resolve().parent
ruta_programacion = list(base_dir.glob("data_raw/*pendientes*.*"))
ruta_actas = list(base_dir.glob("data_raw/*Acta_Clientes*.*"))
ruta_fenix_ans = base_dir / "data_clean" / "FENIX_ANS.xlsx"
ruta_repo = base_dir / "data_clean" / "REPOSITORIO_PEDIDOS_CERRADOS.xlsx"

print("------------------------------------------------------------")
print("🔄 INICIANDO CRUCE PROGRAMACIÓN VS ACTAS")
print("------------------------------------------------------------")

if not ruta_programacion or not ruta_actas:
    print("⚠️ No se encontraron archivos pendientes o actas en data_raw.")
    exit(1)

archivo_prog = max(ruta_programacion, key=lambda f: f.stat().st_mtime)
archivo_actas = max(ruta_actas, key=lambda f: f.stat().st_mtime)
print(f"📘 Programación: {archivo_prog.name}")
print(f"📗 Actas: {archivo_actas.name}")

# ------------------------------------------------------------
# 🧮 LECTOR UNIVERSAL
# ------------------------------------------------------------
def leer_archivo(ruta):
    ext = ruta.suffix.lower()
    if ext in [".csv", ".txt"]:
        try:
            with open(ruta, "r", encoding="utf-8", errors="ignore") as f:
                primera = f.readline()
            if "|" in primera:
                sep = "|"
            elif ";" in primera:
                sep = ";"
            else:
                sep = ","
            df = pd.read_csv(ruta, sep=sep, dtype=str, encoding="utf-8", on_bad_lines="skip")
        except Exception:
            df = pd.read_csv(ruta, sep=sep, dtype=str, encoding="latin1", on_bad_lines="skip")
    elif ext in [".xlsx", ".xls"]:
        df = pd.read_excel(ruta, dtype=str)
    else:
        raise ValueError(f"❌ Tipo de archivo no soportado: {ruta.name}")
    return df

# ------------------------------------------------------------
# 🧩 CARGAR ARCHIVOS
# ------------------------------------------------------------
df_prog = leer_archivo(archivo_prog)
df_actas = leer_archivo(archivo_actas)
df_fenix = pd.read_excel(ruta_fenix_ans, sheet_name="FENIX_ANS", dtype=str)

for df in [df_prog, df_actas, df_fenix]:
    df.columns = df.columns.str.strip().str.lower()

# ------------------------------------------------------------
# 🧩 CRUCE DE PEDIDOS
# ------------------------------------------------------------
pedidos_cumplidos = set(df_actas["pedido"].dropna().unique())
df_prog["estado_cruce"] = df_prog["pedido"].apply(
    lambda x: "CUMPLIDO" if x in pedidos_cumplidos else "PENDIENTE"
)

# ------------------------------------------------------------
# 🔗 (DESACTIVADO) ACTUALIZAR FENIX_ANS con ESTADO_FENIX
# ------------------------------------------------------------
if "pedido" not in df_fenix.columns:
    print("⚠️ No se encontró columna 'pedido' en FENIX_ANS.xlsx. No se puede continuar.")
    exit(1)

print("ℹ️ Cruce Programación vs Actas: se omite actualización de ESTADO_FENIX (columna retirada).")

# ------------------------------------------------------------
# 📦 MOVER PEDIDOS CERRADOS AL REPOSITORIO (flujo limpio)
# ------------------------------------------------------------
print("🔍 Verificando coincidencias antes de mover al repositorio...")

df_fenix_actualizado = pd.read_excel(ruta_fenix_ans, sheet_name="FENIX_ANS", dtype=str)
df_fenix_actualizado.columns = df_fenix_actualizado.columns.str.strip().str.lower()

# 🔒 BLINDAJE: normalizar columnas (evita KeyError por mayúsculas/espacios)
df_fenix_actualizado.columns = (
    df_fenix_actualizado.columns.astype(str).str.strip().str.lower()
)

# 🔎 Buscar pedidos ejecutados en campo y cumplidos (según ACTAS)
df_fenix_actualizado["pedido"] = df_fenix_actualizado["pedido"].astype(str).str.strip()

cerrados = df_fenix_actualizado[
    (df_fenix_actualizado["reporte_tecnico"].astype(str).str.upper().str.contains("EJECUTADO", na=False))
    & (df_fenix_actualizado["pedido"].isin(pedidos_cumplidos))
].copy()

if not cerrados.empty:
    print(f"📦 {len(cerrados)} pedidos cerrados serán movidos al repositorio...")

    # Normalizar columnas
    cerrados.columns = cerrados.columns.str.strip().str.lower()
    cerrados = cerrados.loc[:, ~cerrados.columns.duplicated()]

    # 📁 Si el repositorio existe, combinar datos
    # 📁 Si el repositorio existe, combinar datos
    if ruta_repo.exists():
        repo = pd.read_excel(ruta_repo, dtype=str)
        repo.columns = repo.columns.str.strip().str.lower()
        repo = repo.loc[:, ~repo.columns.duplicated()]

        # 🧹 LIMPIEZA ESPECIAL: eliminar filas vacías y encabezados viejos
        repo = repo[repo.notna().any(axis=1)]
        repo = repo[~repo.apply(lambda fila: any(str(x).strip().lower() in repo.columns for x in fila.values), axis=1)]
        repo.reset_index(drop=True, inplace=True)

        # 🔹 Combinar datos sin desordenar columnas y sin filas vacías
        columnas_repo = list(repo.columns)
        columnas_nuevas = [col for col in cerrados.columns if col not in columnas_repo]

        repo = repo.reindex(columns=columnas_repo + columnas_nuevas)
        cerrados = cerrados.reindex(columns=repo.columns)

       # 🧹 LIMPIEZA ROBUSTA PARA EVITAR FILA VACÍA
        repo.replace(["nan", "None", None], "", inplace=True)

        # Elimina filas que sean 100% vacías o solo espacios
        repo = repo[repo.apply(lambda fila: ''.join(str(v).strip() for v in fila.values) != '', axis=1)]

        repo.reset_index(drop=True, inplace=True)


    else:
        repo = cerrados.copy()

    # 🧹 Limpieza final del repositorio
    repo.drop_duplicates(subset=["pedido"], keep="last", inplace=True)
    repo.dropna(axis=1, how="all", inplace=True)
    # 💾 GUARDAR REPOSITORIO SIN FILA VACÍA
       

    # 💾 GUARDAR REPOSITORIO SIN FILA VACÍA

    # 1️⃣ Crear archivo completamente vacío SIN hoja inicial
    wb = Workbook()
    ws = wb.active
    wb.remove(ws)  # ← EL PASO CLAVE: eliminar la hoja vacía que crea openpyxl
    wb.create_sheet("REPOSITORIO_CERRADOS")
    wb.save(ruta_repo)

    # 2️⃣ Concatenar cerrados ANTES de limpiar
    repo = pd.concat([repo, cerrados], ignore_index=True)

    # 3️⃣ Limpieza robusta (eliminar filas realmente vacías)
    repo.replace(["nan", "None", None], "", inplace=True)
    repo = repo[repo.apply(lambda fila: ''.join(str(v).strip() for v in fila.values) != '', axis=1)]
    repo.drop_duplicates(subset=["pedido"], keep="last", inplace=True)
    repo.reset_index(drop=True, inplace=True)

    # 4️⃣ Guardar de forma limpia sin dejar filas fantasma
    with pd.ExcelWriter(ruta_repo, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        repo.to_excel(writer, sheet_name="REPOSITORIO_CERRADOS", index=False)

    print(f"💾 Archivo actualizado: {ruta_repo}")
  
    # 🧹 Eliminar pedidos cerrados del archivo FENIX_ANS
    print("🧹 Eliminando filas cerradas directamente en FENIX_ANS...")
    wb = load_workbook(ruta_fenix_ans)
    ws = wb["FENIX_ANS"]

    # Detectar columna PEDIDO
    col_pedido = None
    for col in range(1, ws.max_column + 1):
        if str(ws.cell(1, col).value).strip().lower() == "pedido":
            col_pedido = col
            break

    pedidos_cerrados = set(cerrados["pedido"].dropna().astype(str))
    filas_eliminadas = 0

    # Eliminar de abajo hacia arriba
    for i in range(ws.max_row, 1, -1):
        pedido_excel = str(ws.cell(i, col_pedido).value).strip()
        if pedido_excel in pedidos_cerrados:
            ws.delete_rows(i, 1)
            filas_eliminadas += 1

    wb.save(ruta_fenix_ans)
    print(f"✅ {filas_eliminadas} filas eliminadas correctamente de FENIX_ANS.")
    print("------------------------------------------------------------")

    # 🔄 Recargar archivo actualizado antes del formato condicional
    df_fenix_actualizado = pd.read_excel(ruta_fenix_ans, sheet_name="FENIX_ANS", dtype=str)
    df_fenix_actualizado.columns = df_fenix_actualizado.columns.str.strip().str.lower()

    # 🔒 BLINDAJE columnas requeridas para evitar KeyError
    if "estado_fenix" not in df_fenix_actualizado.columns:
        df_fenix_actualizado["estado_fenix"] = ""
        print("⚠️ 'ESTADO_FENIX' no existe en el Excel. Se creó vacía para evitar fallo.")

    if "reporte_tecnico" not in df_fenix_actualizado.columns:
        df_fenix_actualizado["reporte_tecnico"] = ""
        print("⚠️ 'REPORTE_TECNICO' no existe en el Excel. Se creó vacía para evitar fallo.")


    else:
        print("ℹ️ No hay pedidos cerrados nuevos para mover al repositorio.")

# ------------------------------------------------------------
# 🎨 FORMATO CONDICIONAL Y LÓGICA DE ESTADOS
# ------------------------------------------------------------
print("🎨 Aplicando formato condicional en FENIX_ANS...")

wb = load_workbook(ruta_fenix_ans)
ws = wb["FENIX_ANS"]

cols = {str(cell.value).strip().upper(): idx + 1 for idx, cell in enumerate(ws[1])}
col_dias = cols.get("DIAS_RESTANTES")
col_reporte = cols.get("REPORTE_TECNICO")
col_estado = cols.get("ESTADO_FENIX")

if col_estado is None or col_dias is None or col_reporte is None:
    print("⚠️ Faltan columnas para pintar estados. Se omite el formato para evitar errores.")
    wb.save(ruta_fenix_ans)
    exit(0)
    
    print(f"🎨 Columnas detectadas correctamente → REPORTE: {col_reporte}, ESTADO: {col_estado}")

# 🎨 Colores
verde = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
naranja = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
rojo = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
gris = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

# 🔄 Aplicar reglas de negocio
for fila in range(2, ws.max_row + 1):
    try:
        # ✅ No tocar los pedidos que ya están cumplidos
        if str(ws.cell(fila, col_estado).value).strip().upper() == "CUMPLIDO":
            continue

        reporte = (ws.cell(fila, col_reporte).value or "").strip().upper()
        dias_texto = str(ws.cell(fila, col_dias).value or "")
        celda_estado = ws.cell(fila, col_estado)

        # 1️⃣ Si el técnico no ha reportado nada:
        if reporte == "SIN DATO" or reporte == "":
            celda_estado.value = "ABIERTO"
            celda_estado.fill = gris
            continue

        # 2️⃣ Si ya está ejecutado en campo:
        if "EJECUTADO" in reporte:
            dias_num = 0
            if "día" in dias_texto:
                try:
                    dias_num = int(dias_texto.split("día")[0].strip())
                except:
                    dias_num = 0

            if dias_num > 2:
                celda_estado.value = "A TIEMPO"
                celda_estado.fill = verde
            elif 0 < dias_num <= 2:
                celda_estado.value = "ALERTA"
                celda_estado.fill = amarillo
            elif dias_num == 0 and "hora" in dias_texto:
                celda_estado.value = "A CERO"
                celda_estado.fill = naranja
            elif dias_num < 0:
                celda_estado.value = "VENCIDO"
                celda_estado.fill = rojo
            else:
                celda_estado.value = "ALERTA"
                celda_estado.fill = amarillo
        else:
            celda_estado.value = "ABIERTO"
            celda_estado.fill = gris

    except Exception as e:
        print(f"⚠️ Error procesando fila {fila}: {e}")

wb.save(ruta_fenix_ans)
print("✅ Formato condicional aplicado correctamente.")
print("------------------------------------------------------------")
print("✅ Cruce, actualización y formatos finalizados.")
print("------------------------------------------------------------")
