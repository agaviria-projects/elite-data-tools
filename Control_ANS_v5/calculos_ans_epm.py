"""
------------------------------------------------------------
CÁLCULOS ANS EPM – Proyecto Control_ANS_FENIX
------------------------------------------------------------
Autor: Héctor + IA (2025)
------------------------------------------------------------
Descripción:
- Lee el archivo limpio (FENIX_CLEAN.xlsx)
- Calcula días pactados, fecha límite ANS, estado y métricas.
- Excluye sábados, domingos y festivos.
- Mantiene hora/minuto del inicio.
- Exporta a FENIX_ANS.xlsx con hoja RESUMEN.
------------------------------------------------------------
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule


# ------------------------------------------------------------
# ⚙️ CONFIGURACIÓN GLOBAL DE ADVERTENCIAS
# ------------------------------------------------------------
import warnings
warnings.filterwarnings("ignore", category=UserWarning)

# ------------------------------------------------------------
# CONFIGURACIÓN DE RUTAS PORTABLES
# ------------------------------------------------------------
base_path = Path(__file__).resolve().parent

ruta_input  = base_path / "data_clean" / "FENIX_CLEAN.xlsx"
ruta_output = base_path / "data_clean" / "FENIX_ANS_EPM.xlsx"


# 👉 ESTA LÍNEA DEBE IR AQUÍ
# ruta_repo   = base_path / "data_clean" / "REPOSITORIO_PEDIDOS_CERRADOS.xlsx"



# ------------------------------------------------------------
# CONFIGURACIÓN DE CALENDARIO SIN DÍAS FESTIVOS
# ------------------------------------------------------------
WEEKMASK = "1111100"  # lunes a viernes

FESTIVOS = np.array([
    "2025-01-01","2025-01-06","2025-03-24","2025-04-17","2025-04-18",
    "2025-05-01","2025-05-26","2025-06-16","2025-06-23","2025-07-07",
    "2025-08-07","2025-08-18","2025-10-13","2025-11-03","2025-11-17",
    "2025-12-08","2025-12-25",

    "2026-01-01","2026-01-12","2026-03-23","2026-04-02","2026-04-03",
    "2026-05-01","2026-05-18","2026-06-08","2026-06-15","2026-06-29",
    "2026-07-13",  # ← Nuevo festivo decretado
    "2026-07-20",
    "2026-08-07","2026-08-17","2026-10-12","2026-11-02",
    "2026-11-16","2026-12-08","2026-12-25"
], dtype="datetime64[D]")

# ------------------------------------------------------------
# FUNCIÓN: FECHA LÍMITE SEGÚN LÓGICA FÉNIX
# ------------------------------------------------------------
def add_business_days_keep_time(start_dt, n_days):
    if pd.isna(start_dt) or n_days <= 0:
        return pd.NaT

    date_part = np.datetime64(start_dt.date())
    time_part = start_dt.time()

    # Día no hábil → primer hábil siguiente
    if not np.is_busday(date_part, weekmask=WEEKMASK, holidays=FESTIVOS):
        primer_habil = np.busday_offset(date_part, 0, roll="forward",
                                        weekmask=WEEKMASK, holidays=FESTIVOS)
        limite = np.busday_offset(primer_habil, n_days - 1, roll="forward",
                                  weekmask=WEEKMASK, holidays=FESTIVOS)
    else:
        # Día hábil → siguiente hábil
        siguiente_habil = np.busday_offset(date_part, 1, roll="forward",
                                           weekmask=WEEKMASK, holidays=FESTIVOS)
        limite = np.busday_offset(siguiente_habil, n_days - 1, roll="forward",
                                  weekmask=WEEKMASK, holidays=FESTIVOS)

    return datetime.combine(pd.to_datetime(str(limite)).date(), time_part)

# ------------------------------------------------------------
# FUNCIÓN: DÍAS HÁBILES ENTRE DOS FECHAS
# ------------------------------------------------------------
def business_days_between(start_dt, end_dt):
    if pd.isna(start_dt) or pd.isna(end_dt):
        return np.nan
    start_date = np.datetime64(start_dt.date() + timedelta(days=1))
    end_date = np.datetime64(end_dt.date())
    dias = np.busday_count(start_date, end_date, weekmask=WEEKMASK, holidays=FESTIVOS)
    if np.is_busday(end_date, weekmask=WEEKMASK, holidays=FESTIVOS) and end_date > start_date:
        dias += 1
    return int(dias)

# ------------------------------------------------------------
# CARGA DE DATOS
# ------------------------------------------------------------
df = pd.read_excel(ruta_input)
print(df[df["PEDIDO"].astype(str).str.contains("2275", na=False)])
print(f"📂 Archivo cargado: {ruta_input.name} ({len(df)} registros)")

# ------------------------------------------------------------
# LIMPIEZA Y CONVERSIÓN DE FECHAS
# ------------------------------------------------------------
columnas_clave = ["PEDIDO", "FECHA_INICIO_ANS", "TIPO_DIRECCION", "ACTIVIDAD"]

from pandas.api.types import is_datetime64_any_dtype

for col in columnas_clave:

    if is_datetime64_any_dtype(df[col]):

        df[col] = df[col].apply(
            lambda x: np.nan if pd.isna(x) else x
        )

    else:

        df[col] = df[col].apply(
            lambda x: np.nan
            if str(x).strip() == ""
            or str(x).upper() in ["NAN", "NONE", "NULL"]
            else x
        )

# Nota: la advertencia "Parsing dates..." es solo informativa y no afecta el flujo.
# Se mantiene 'dayfirst=True' para compatibilidad con formatos DD/MM/YYYY y YYYY/MM/DD.
df["FECHA_INICIO_ANS"] = pd.to_datetime(df["FECHA_INICIO_ANS"], errors="coerce", dayfirst=True)

# ------------------------------------------------------------
# DÍAS EPM
# ------------------------------------------------------------
DIAS_PACTADOS_MAP = {
    "ACREV":  {"URBANO": 5,  "RURAL": 5},
    "ALEGN":  {"URBANO": 8,  "RURAL": 11},
    "ALEGA":  {"URBANO": 8,  "RURAL": 11},
    "ALECA":  {"URBANO": 8,  "RURAL": 11},
    "ACAMN":  {"URBANO": 8,  "RURAL": 11},
    "AMRTR":  {"URBANO": 10, "RURAL": 15},
    "REEQU":  {"URBANO": 14, "RURAL": 14},
    "INPRE":  {"URBANO": 14, "RURAL": 14},
    "DIPRE":  {"URBANO": 14, "RURAL": 14},
    "ARTER":  {"URBANO": 6,  "RURAL": 10},
    "AEJDO":  {"URBANO": 9,  "RURAL": 14},
    "VITEC":  {"URBANO": 2,  "RURAL": 2},
    "APLIN":  {"URBANO": 9,  "RURAL": 9},
}

def dias_pactados(row):
    act = str(row.get("ACTIVIDAD", "")).strip().upper()
    tipo = str(row.get("TIPO_DIRECCION", "")).strip().upper()
    if act in DIAS_PACTADOS_MAP and tipo in DIAS_PACTADOS_MAP[act]:
        return DIAS_PACTADOS_MAP[act][tipo]
    return 0

df["DIAS_PACTADOS"] = df.apply(dias_pactados, axis=1)

# ------------------------------------------------------------
# FECHA LÍMITE ANS
# ------------------------------------------------------------
df["FECHA_LIMITE_ANS"] = df.apply(
    lambda r: add_business_days_keep_time(r["FECHA_INICIO_ANS"], r["DIAS_PACTADOS"]),
    axis=1
)

# ------------------------------------------------------------
# DÍAS TRANSCURRIDOS
# ------------------------------------------------------------
hoy = datetime.now()

def ajustar_hora(fecha_inicio):
    if pd.isna(fecha_inicio):
        return hoy
    return hoy.replace(hour=fecha_inicio.hour, minute=fecha_inicio.minute, second=fecha_inicio.second, microsecond=0)

def calcular_dias_transcurridos(row):
    fecha_ini = row["FECHA_INICIO_ANS"]
    if pd.isna(fecha_ini):
        return ""
    hoy_ref = ajustar_hora(fecha_ini)
    dias_habiles = business_days_between(fecha_ini, hoy_ref)
    hora_inicio = fecha_ini.strftime("%H:%M")
    return f"{dias_habiles} días {hora_inicio}"

df["DIAS_TRANSCURRIDOS"] = df.apply(calcular_dias_transcurridos, axis=1)

# ------------------------------------------------------------
# DÍAS RESTANTES (ajuste exacto incluyendo fin de semana y hora)
# ------------------------------------------------------------
def calcular_dias_restantes(row):
    fecha_lim = row["FECHA_LIMITE_ANS"]
    fecha_ini = row["FECHA_INICIO_ANS"]
    if pd.isna(fecha_lim) or pd.isna(fecha_ini):
        return ""

    hoy = datetime.now()
    hora_ref = fecha_ini.time()

    # Si ya venció
    if hoy > fecha_lim:
        return "VENCIDO"

    # Calcular días hábiles restantes sin sumar extra
    dias_habiles = np.busday_count(
        np.datetime64(hoy.date()),
        np.datetime64(fecha_lim.date()),
        weekmask=WEEKMASK,
        holidays=FESTIVOS
    )

    # ✅ Ajuste: si el siguiente día hábil es el mismo del límite, poner 1 día
    if dias_habiles == 0 and hoy.date() != fecha_lim.date():
        dias_habiles = 1

    # Si el día límite es hoy
    if hoy.date() == fecha_lim.date():
        if hoy.time() < fecha_lim.time():
            return f"0 días {fecha_ini.strftime('%H:%M')}"
        else:
            return "VENCIDO"

    # Si hoy es viernes y el vencimiento es lunes (fin de semana de por medio)
    # => contar solo el lunes como 1 día
    siguiente_habil = np.busday_offset(
        np.datetime64(hoy.date()), 1, roll="forward", weekmask=WEEKMASK, holidays=FESTIVOS
    )
    if siguiente_habil == np.datetime64(fecha_lim.date()):
        dias_habiles = 1

    return f"{dias_habiles} días {fecha_ini.strftime('%H:%M')}"

df["DIAS_RESTANTES"] = df.apply(calcular_dias_restantes, axis=1)

# ------------------------------------------------------------
# ESTADO
# ------------------------------------------------------------
def calcular_estado(row):

    fecha_lim = row["FECHA_LIMITE_ANS"]

    if pd.isna(fecha_lim):
        return "SIN FECHA"

    ahora = datetime.now()

    if ahora >= fecha_lim:
        return "VENCIDO"

    if ahora.date() == fecha_lim.date():
        return "ALERTA_0 Días"

    dias_restantes = np.busday_count(
        np.datetime64(ahora.date()),
        np.datetime64(fecha_lim.date()),
        weekmask=WEEKMASK,
        holidays=FESTIVOS
    )

    if dias_restantes <= 2:
        return "ALERTA"

    return "A TIEMPO"

# ------------------------------------------------------------
# GENERAR ESTADO
# ------------------------------------------------------------
df["ESTADO"] = df.apply(calcular_estado, axis=1)

print("✅ Estado ANS calculado correctamente")
# ------------------------------------------------------------
# VERIFICAR SI EL ARCHIVO FENIX_ANS ESTÁ ABIERTO
# ------------------------------------------------------------
import os
import tkinter as tk
from tkinter import messagebox

def verificar_archivo_abierto(ruta):
    """Verifica si el archivo Excel está en uso por Excel u otro proceso."""
    if os.path.exists(ruta):
        try:
            with open(ruta, "a"):
                pass  # Si puede abrirse, no está bloqueado
        except PermissionError:
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror(
                "🚫 Archivo bloqueado",
                "El Informe' está abierto en Excel.\n\n"
                "🔒 Cierra el archivo y vuelve a ejecutar el proceso."
            )
            print("⛔ Proceso detenido: el archivo está abierto.")
            exit()
# ------------------------------------------------------------
# 🔗 CRUCE CON GOOGLE SHEETS – FORMULARIO CONTROL ANS (v8.1 blindado)
# ------------------------------------------------------------
import gspread
from google.oauth2.service_account import Credentials
import re

def limpiar_pedido(x):
    """
    Limpia cualquier valor de PEDIDO:
    - Convierte a texto
    - Elimina espacios invisibles
    - Quita ceros adelante
    - Elimina .0 si viene como número flotante
    """
    if pd.isna(x):
        return ""
    x = str(x).strip()

    # Quitar caracteres invisibles
    x = re.sub(r"[\u200B-\u200D\uFEFF\u00A0]", "", x)

    # Quitar .0 de Excel
    if x.endswith(".0"):
        x = x.replace(".0", "")

    # Quitar ceros a la izquierda
    x = x.lstrip("0")

    return x


try:
    cred_path = base_path / "control-ans-elite-f4ea102db569.json"  # <--- CORRECTO
    scopes = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
    creds = Credentials.from_service_account_file(cred_path, scopes=scopes)
    client = gspread.authorize(creds)


    SHEET_ID = "1bPLGVVz50k6PlNp382isJrqtW_3IsrrhGW0UUlMf-bM"
    sheet = client.open_by_key(SHEET_ID)

    hoja = None
    for ws in sheet.worksheets():
        if "RESP" in ws.title.upper() or "FORM" in ws.title.upper():
            hoja = ws
            break

    if hoja is None:
        raise Exception("No se detectó pestaña válida del formulario.")

    data = hoja.get_all_records()

    if not data:
        print("⚠️ Formulario vacío — se dejan columnas en SIN DATO.")
        df["REPORTE_TECNICO"] = "SIN DATO"
        df["TECNICO_EJECUTA"] = "SIN DATO"
    else:
        df_form = pd.DataFrame(data)
        df_form.rename(columns=lambda c: c.strip().upper(), inplace=True)

        # Renombrar columnas
        renames = {
            "NÚMERO DEL PEDIDO": "PEDIDO",
            "ESTADO DEL PEDIDO": "REPORTE_TECNICO",
            "NOMBRE DEL TÉCNICO": "TECNICO_EJECUTA",
            "OBSERVACIÓN": "OBSERVACION"
        }
        df_form.rename(columns=renames, inplace=True)

        # Normalizar pedidos
        df["PEDIDO"] = df["PEDIDO"].apply(limpiar_pedido)
        df_form["PEDIDO"] = df_form["PEDIDO"].apply(limpiar_pedido)

        # Limpiar textos del formulario
        if "REPORTE_TECNICO" in df_form.columns:
            df_form["REPORTE_TECNICO"] = df_form["REPORTE_TECNICO"].astype(str).str.upper().str.strip()

        if "TECNICO_EJECUTA" in df_form.columns:
            df_form["TECNICO_EJECUTA"] = df_form["TECNICO_EJECUTA"].astype(str).str.upper().str.strip()

        # MERGE SEGURO
        columnas = ["PEDIDO", "REPORTE_TECNICO", "TECNICO_EJECUTA","OBSERVACION"]
        columnas = [c for c in columnas if c in df_form.columns]

        df = df.merge(df_form[columnas], on="PEDIDO", how="left")

        print(
            df.groupby("PEDIDO")
            .size()
            .reset_index(name="conteo")
            .query("conteo > 1")
            .head(10)
        )

        # Rellenar vacíos
        df["REPORTE_TECNICO"] = df["REPORTE_TECNICO"].fillna("SIN DATO")
        df["TECNICO_EJECUTA"] = df["TECNICO_EJECUTA"].fillna("SIN DATO")
        df["OBSERVACION"] = df["OBSERVACION"].fillna("SIN DATO")
       
        # ============================================================
        # 🧹 LIMPIEZA INTELIGENTE DE DUPLICADOS "SIN DATO"
        # ============================================================

        # Identificar filas totalmente vacías funcionalmente
        mask_sin_dato = (
            (df["REPORTE_TECNICO"] == "SIN DATO") &
            (df["TECNICO_EJECUTA"] == "SIN DATO") &
            (df["OBSERVACION"] == "SIN DATO")
        )

        # Separar
        df_con_info = df[~mask_sin_dato].copy()
        df_sin_info = df[mask_sin_dato].copy()

        # Colapsar SOLO los SIN DATO por PEDIDO
        df_sin_info = df_sin_info.drop_duplicates(subset=["PEDIDO"], keep="first")

        # Reconstruir dataframe final
        df = pd.concat([df_con_info, df_sin_info], ignore_index=True)

        print("🧹 Duplicados SIN DATO colapsados correctamente")    
        
        print("🔗 Cruce con Google Sheets finalizado correctamente ✔")

except Exception as e:
    print(f"⚠️ Error en cruce con formulario Google Sheets: {e}")
    df["REPORTE_TECNICO"] = df.get("REPORTE_TECNICO", "SIN DATO")
    df["TECNICO_EJECUTA"] = df.get("TECNICO_EJECUTA", "SIN DATO")


# ============================================================
# 🩹 CREAR COLUMNAS SI NO EXISTEN (solución definitiva)
# ============================================================
for columna in ["REPORTE_TECNICO", "TECNICO_EJECUTA", "ESTADO_FENIX","OBSERVACION"]:
    if columna not in df.columns:
        df[columna] = "SIN DATO"
        print(f"🆕 Columna agregada automáticamente: {columna}")


        # 🔧 Crear columnas necesarias si Google Sheets falla
    if "REPORTE_TECNICO" not in df.columns:
        df["REPORTE_TECNICO"] = "SIN DATO"

    if "TECNICO_EJECUTA" not in df.columns:
        df["TECNICO_EJECUTA"] = "SIN DATO"


# # # ------------------------------------------------------------
# # # 🧭 NUEVA COLUMNA: ESTADO_FENIX (según cruce FENIX + formulario)
# # # ------------------------------------------------------------
"""
🔒 BLOQUE DESACTIVADO (v4.1)
Esta sección se comenta nuevamente después de generar el informe ANS.
El cálculo real de ESTADO_FENIX se realizará mediante el script:
➡️ cruce_digitacion_fenix.py
que usa Digitacion Fenix.txt como fuente oficial.
"""
# from datetime import datetime
# import pandas as pd

# hoy = datetime.now()

# def calcular_estado_fenix(row):
#     form = str(row.get("REPORTE_TECNICO", "")).strip().upper()
#     estado_fenix_origen = str(row.get("ESTADO_FENIX_ORIGEN", "")).strip().upper()
#     fecha_lim = pd.to_datetime(row.get("FECHA_LIMITE_ANS", ""), errors="coerce")

#     if pd.isna(fecha_lim):
#         return "SIN FECHA"

#     dias_rest = (fecha_lim.date() - hoy.date()).days

#     # ✅ CERRADO solo si el técnico ejecutó en campo y FENIX lo confirma cerrado
#     if form == "EJECUTADO EN CAMPO" and estado_fenix_origen == "CERRADO":
#         return "CERRADO"

#     # 🟡 Si el técnico ejecutó pero FENIX aún no está cerrado
#     if form == "EJECUTADO EN CAMPO" and estado_fenix_origen != "CERRADO":
#         return "PENDIENTE VALIDACIÓN"

#     # 🔴 Semáforo ANS (según fecha límite)
#     if dias_rest < 0:
#         return "VENCIDO"
#     elif dias_rest == 0:
#         return "CRÍTICO"
#     elif dias_rest < 2:
#         return "APUNTO DE VENCER"
#     else:
#         return "ABIERTO"

# df["ESTADO_FENIX"] = df.apply(calcular_estado_fenix, axis=1)
# print("🧭 Columna ESTADO_FENIX generada correctamente con validación cruzada.")

# ------------------------------------------------------------
# 📦 MOVER PEDIDOS CUMPLIDOS A REPOSITORIO HISTÓRICO (VERSIÓN FINAL)
# Condición:
#   - REPORTE_TECNICO = LEGALIZADO
#   - ESTADO_FENIX   = CUMPLIDO
# El repositorio conserva TODAS las columnas de FENIX_ANS
# ------------------------------------------------------------

# Normalizar columnas clave
# df["PEDIDO"] = df["PEDIDO"].astype(str).str.strip()
# df["REPORTE_TECNICO"] = df["REPORTE_TECNICO"].astype(str).str.upper().str.strip()
# df["ESTADO_FENIX"] = df["ESTADO_FENIX"].astype(str).str.upper().str.strip()

# # Filtrar pedidos que cumplen ambas condiciones
# cerrados = df[
#     (df["REPORTE_TECNICO"] == "LEGALIZADO") &
#     (df["ESTADO_FENIX"] == "CUMPLIDO")
# ].copy()

# if cerrados.empty:
#     print("ℹ️ No se encontraron pedidos LEGALIZADOS y CUMPLIDOS para mover.")
# else:
#     print(f"📦 {len(cerrados)} pedidos serán movidos al repositorio histórico.")

#     # Si el repositorio existe, cargarlo
#     if ruta_repo.exists():
#         repo = pd.read_excel(ruta_repo, dtype=str)

#         # Asegurar que tenga TODAS las columnas actuales
#         for col in df.columns:
#             if col not in repo.columns:
#                 repo[col] = ""

#         # Alinear exactamente el orden de columnas
#         repo = repo[df.columns]

#         # Concatenar
#         repo = pd.concat([repo, cerrados], ignore_index=True)

#         # Eliminar duplicados por PEDIDO (conservar el más reciente)
#         repo["PEDIDO"] = repo["PEDIDO"].astype(str).str.strip()
#         repo = repo.drop_duplicates(subset=["PEDIDO"], keep="last")

#     else:
#         # Crear repositorio nuevo con estructura completa
#         repo = cerrados.copy()

#     # Guardar repositorio
#     repo.to_excel(ruta_repo, index=False)
#     print("🗂️ Repositorio histórico actualizado correctamente.")

#     # Eliminar pedidos movidos del archivo principal
#     df = df[~df["PEDIDO"].isin(cerrados["PEDIDO"])].copy()
#     print("🧹 Pedidos eliminados de FENIX_ANS tras ser archivados.")

# ------------------------------------------------------------
# 🔍 CRUCE PARA INSERTAR COORDENADAS Y ZONAS (Z – AC) – CONSOLIDADO
# ------------------------------------------------------------

archivos_pend = sorted(
    (base_path / "data_raw").glob("pendientes_*.csv"),
    key=lambda x: x.stat().st_mtime
)

if not archivos_pend:
    print("⚠️ No se encontró ningún archivo pendientes_* para cargar coordenadas.")
else:
    print("📂 Archivos de pendientes usados para coordenadas:")
    dfs_pen = []

    for archivo in archivos_pend:
        print(f"   - {archivo.name}")
        df_tmp = pd.read_csv(archivo, dtype=str, encoding="latin-1")
        df_tmp.columns = df_tmp.columns.str.strip().str.upper()
        dfs_pen.append(df_tmp)

    # Consolidar TODOS los pendientes
    df_pen = pd.concat(dfs_pen, ignore_index=True)

    # Columnas requeridas
    columnas_req = ["PEDIDO", "COORDENADAX", "COORDENADAY", "AREA_OPERATIVA", "SUBZONA", "SUBPED"]
    faltantes = [c for c in columnas_req if c not in df_pen.columns]

    if faltantes:
        print(f"⚠️ Columnas faltantes en pendientes: {faltantes}")
    else:
        # Normalizar pedido
        df["PEDIDO"] = df["PEDIDO"].astype(str).str.strip()
        df_pen["PEDIDO"] = df_pen["PEDIDO"].astype(str).str.strip()

        # DataFrame auxiliar solo con lo necesario
        df_merge = df_pen[columnas_req].copy()

        # Merge seguro
        df = df.merge(df_merge, on="PEDIDO", how="left")

        # ------------------------------------------------------------
        # 🎯 FILTRO DEFINITIVO DE SUBZONAS VÁLIDAS ANS
        # ------------------------------------------------------------
        # ⚠️ IMPORTANTE:
        # Si FÉNIX agrega nuevas zonas/subzonas,
        # deben incluirse explícitamente en esta lista.
        # El mapa y Excel se actualizan automáticamente.
        subzonas_validas = [
            "METROPOLITANA SUR",
            "OCCIDENTE",
            "SUROESTE",
            "NORDESTE",
            "ORIENTE"
        ]

        # Normalizar texto
        df["SUBZONA"] = (
            df["SUBZONA"]
            .astype(str)
            .str.upper()
            .str.strip()
        )

        # Aplicar filtro
        df = df[df["SUBZONA"].isin(subzonas_validas)]

        # ------------------------------------------------------------
        # 🏙️ AJUSTE TERRITORIAL TIPO_DIRECCION
        # ------------------------------------------------------------
        # Regla de negocio:
        # Para SUBZONA SUROESTE y OCCIDENTE,METROPOLITANA SUR
        # instalaciones que inician con "190" y cuyo 7.º dígito es "1"
        # se consideran URBANAS según criterio operativo (2025)

        mask_urbano_especial = (
            df["SUBZONA"].isin(["SUROESTE", "OCCIDENTE","METROPOLITANA SUR","NORDESTE","ORIENTE"]) &
            df["INSTALACION"].astype(str).str.startswith("190") &
            df["INSTALACION"].astype(str).str.len().ge(7) &
            (df["INSTALACION"].astype(str).str[6] == "1")
        )

        df.loc[mask_urbano_especial, "TIPO_DIRECCION"] = "URBANO"
        
        
        # 🔁 RECÁLCULO NECESARIO TRAS AJUSTE TERRITORIAL
        df["DIAS_PACTADOS"] = df.apply(dias_pactados, axis=1)

        df["FECHA_LIMITE_ANS"] = df.apply(
            lambda r: add_business_days_keep_time(
                r["FECHA_INICIO_ANS"], r["DIAS_PACTADOS"]
            ),
            axis=1
        )

        # Convertir coordenadas a numérico
        df["COORDENADAX"] = pd.to_numeric(df["COORDENADAX"], errors="coerce")
        df["COORDENADAY"] = pd.to_numeric(df["COORDENADAY"], errors="coerce")

        print("📍 Coordenadas, zonas y subzonas consolidadas correctamente desde TODOS los CSV.")

        # ------------------------------------------------------------
        # RECALCULAR SEMÁFORO FINAL
        # ------------------------------------------------------------
        df["DIAS_RESTANTES"] = df.apply(
            calcular_dias_restantes,
            axis=1
        )

        df["ESTADO"] = df.apply(
            calcular_estado,
            axis=1
        )

        
# ============================================================
# 🔒 COLAPSO DEFINITIVO DE DUPLICADOS
# REGLA DE NEGOCIO: 1 PEDIDO = 1 FILA
# ============================================================

print("🔒 Colapsando pedidos duplicados (regla negocio: 1 pedido = 1 fila)")

# Prioridad:
# 1. Pedidos con información (LEGALIZADO / EJECUTADO / con observación)
# 2. Pedidos SIN DATO quedan al final
df = (
    df
    .sort_values(
        by=["PEDIDO", "REPORTE_TECNICO", "OBSERVACION"],
        ascending=[True, False, False]
    )
    .drop_duplicates(subset=["PEDIDO"], keep="first")
    .reset_index(drop=True)
)

print("✅ Duplicados eliminados definitivamente por PEDIDO")

print("📊 TOTAL PEDIDOS ÚNICOS POST-COLAPSO:", df["PEDIDO"].nunique())
print("📍 Subzonas finales POST-COLAPSO (DATASET FINAL):")
print(df["SUBZONA"].value_counts())
# ------------------------------------------------------------
# EXPORTAR ARCHIVO
# ------------------------------------------------------------
verificar_archivo_abierto(ruta_output)  # 👈 ESTA LÍNEA ES CLAVE
# ------------------------------------------------------------
# 🔧 NORMALIZAR FECHAS PARA EVITAR DESFASES EN POWER BI
# ------------------------------------------------------------
# Se exportan como texto plano ISO (no tipo datetime)
# Así Power BI las lee exactamente igual sin conversión de zona ni AM/PM

df["FECHA_INICIO_ANS"] = df["FECHA_INICIO_ANS"].apply(
    lambda x: x.strftime("%Y-%m-%d %H:%M:%S") if pd.notnull(x) else ""
)
df["FECHA_LIMITE_ANS"] = df["FECHA_LIMITE_ANS"].apply(
    lambda x: x.strftime("%Y-%m-%d %H:%M:%S") if pd.notnull(x) else ""
)


ruta_output.parent.mkdir(exist_ok=True)
with pd.ExcelWriter(ruta_output, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="FENIX_ANS")
    resumen = df["ESTADO"].value_counts().reset_index()
    resumen.columns = ["ESTADO", "CANTIDAD"]
    resumen.to_excel(writer, index=False, sheet_name="RESUMEN")

print("✅ Cálculos ANS completados correctamente.")
print(f"📁 Archivo exportado: {ruta_output}")


# ============================================================
# 🔒 FORMATO FINAL CONSOLIDADO – UN SOLO load_workbook / save
# ============================================================

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime

wb = load_workbook(ruta_output)
ws = wb["FENIX_ANS"]

ultima_fila = ws.max_row
ultima_col = ws.max_column
ultima_col_letra = ws.cell(row=1, column=ultima_col).column_letter


# ------------------------------------------------------------
# 🔧 FUNCIÓN SEGURA: OBTENER LETRA DE COLUMNA POR NOMBRE
# ------------------------------------------------------------
def obtener_columna_por_nombre(ws, nombre_col):
    for celda in ws[1]:
        if str(celda.value).strip().upper() == nombre_col.upper():
            return celda.column_letter
    raise ValueError(f"Columna '{nombre_col}' no encontrada en Excel")


# ------------------------------------------------------------
# 🎯 OBTENER COLUMNAS DINÁMICAS (CRÍTICO)
# ------------------------------------------------------------
col_estado = obtener_columna_por_nombre(ws, "ESTADO")
col_form   = obtener_columna_por_nombre(ws, "REPORTE_TECNICO")
try:
    col_fenix = obtener_columna_por_nombre(ws, "ESTADO_FENIX")
except ValueError:
    col_fenix = None


# ============================================================
# 🎨 FORMATO CONDICIONAL – ESTADO
# ============================================================
rango_estado = f"${col_estado}$2:${col_estado}${ultima_fila}"

ws.conditional_formatting.add(
    rango_estado,
    FormulaRule(
        formula=[f'${col_estado}2="VENCIDO"'],
        fill=PatternFill(fill_type="solid", start_color="FF0000", end_color="FF0000")
    )
)

ws.conditional_formatting.add(
    rango_estado,
    FormulaRule(
        formula=[f'${col_estado}2="ALERTA_0 Días"'],
        fill=PatternFill(fill_type="solid", start_color="FFA500", end_color="FFA500")
    )
)

ws.conditional_formatting.add(
    rango_estado,
    FormulaRule(
        formula=[f'${col_estado}2="ALERTA"'],
        fill=PatternFill(fill_type="solid", start_color="FFF200", end_color="FFF200")
    )
)

ws.conditional_formatting.add(
    rango_estado,
    FormulaRule(
        formula=[f'${col_estado}2="A TIEMPO"'],
        fill=PatternFill(fill_type="solid", start_color="00B050", end_color="00B050")
    )
)


# ============================================================
# 🎨 FORMATO CONDICIONAL – REPORTE_TECNICO
# ============================================================
rango_form = f"${col_form}$2:${col_form}${ultima_fila}"

ws.conditional_formatting.add(
    rango_form,
    FormulaRule(
        formula=[f'EXACT(${col_form}2,"SIN DATO")'],
        fill=PatternFill(fill_type="solid", start_color="D9D9D9", end_color="D9D9D9")
    )
)


# ============================================================
# 🎨 FORMATO CONDICIONAL – ESTADO_FENIX
# ============================================================
# rango_fenix = f"${col_fenix}$2:${col_fenix}${ultima_fila}"

# ws.conditional_formatting.add(
#     rango_fenix,
#     FormulaRule(
#         formula=[f'${col_fenix}2="CERRADO"'],
#         fill=PatternFill(fill_type="solid", start_color="00B050", end_color="00B050"),
#         font=Font(color="FFFFFF")
#     )
# )

# ws.conditional_formatting.add(
#     rango_fenix,
#     FormulaRule(
#         formula=[f'${col_fenix}2="VENCIDO"'],
#         fill=PatternFill(fill_type="solid", start_color="FF0000", end_color="FF0000"),
#         font=Font(color="FFFFFF")
#     )
# )


# ============================================================
# 💄 TABLA ESTRUCTURADA SEGURA (NO DUPLICA)
# ============================================================
if "FENIX_ANS_TABLA" not in ws.tables:
    tabla = Table(
        displayName="FENIX_ANS_TABLA",
        ref=f"A1:{ultima_col_letra}{ultima_fila}"
    )
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True
    )
    ws.add_table(tabla)


# ============================================================
# 💡 AJUSTES VISUALES
# ============================================================
ws.sheet_view.showGridLines = False

for col in ws.columns:
    max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    ws.column_dimensions[col[0].column_letter].width = max_len + 2

for col_name in ["I", "J"]:
    for cell in ws[col_name]:
        cell.alignment = Alignment(horizontal="center")


# ============================================================
# 📋 HOJA CONFIG_DIAS_PACTADOS
# ============================================================
if "CONFIG_DIAS_PACTADOS" in wb.sheetnames:
    del wb["CONFIG_DIAS_PACTADOS"]

ws_conf = wb.create_sheet("CONFIG_DIAS_PACTADOS")
ws_conf.append(["Actividad", "Descripción", "Días Urbanos", "Días Rurales"])

datos = [
    ["ACREV","PUNTOS CONEXIÓN",5,5],
    ["ALEGN","LEGALIZACIÓN",8,11],
    ["ALEGA","LEGALIZACIÓN",8,11],
    ["ALECA","LEGALIZACIÓN",8,11],
    ["ACAMN","REFORMA",8,11],
    ["AMRTR","MOVIMIENTO RED",10,15],
    ["REEQU","PREPAGO",14,14],
    ["INPRE","INSTALACIÓN",14,14],
    ["DIPRE","DESINSTALAR",14,14],
    ["ARTER","REPLANTEO",6,10],
    ["AEJDO","EJECUCIÓN",9,14],
    ["VITEC","VITEC",2,2],
    ["APLIN","APLIN",9,9],
]


for f in datos:
    ws_conf.append(f)


# ============================================================
# 📋 HOJA META_INFO
# ============================================================
if "META_INFO" in wb.sheetnames:
    del wb["META_INFO"]

ws_meta = wb.create_sheet("META_INFO")
ws_meta["A1"] = "Fuente"
ws_meta["B1"] = "FENIX"
ws_meta["A2"] = "Fecha proceso"
ws_meta["B2"] = datetime.now().strftime("%d/%m/%Y %H:%M")


# ============================================================
# 💾 GUARDADO FINAL ÚNICO (CRÍTICO)
# ============================================================
wb.save(ruta_output)
print("✅ FENIX_ANS_EPM.xlsx generado correctamente (archivo estable, sin corrupción).")