"""
------------------------------------------------------------
VALIDACIÓN DE PEDIDOS DIGITADOS (ALMACÉN VS FENIX)
------------------------------------------------------------
Autor: Héctor + IA (2025)
------------------------------------------------------------
Descripción:
- Cruza FENIX_ANS (programación) vs ALMACEN_EXPORT (digitación)
- Verifica materiales obligatorios por mano de obra.
- Genera VALIDACION_EXPORT.xlsx con:
  ✅ Tabla estructurada
  🟢 Amarillo 🔴 Iconos de estado visuales.
------------------------------------------------------------
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule

# ------------------------------------------------------------
# RUTAS
# ------------------------------------------------------------
base = Path(__file__).resolve().parent
ruta_fenix = base / "data_clean" / "FENIX_ANS.xlsx"
ruta_almacen = base / "data_raw" / "ALMACEN_EXPORT.xlsx"
ruta_relacion = base / "data_raw" / "RELACION_MO_MAT.xlsx"
ruta_salida = base / "data_clean" / "VALIDACION_EXPORT.xlsx"

# ------------------------------------------------------------
# CARGA DE ARCHIVOS
# ------------------------------------------------------------
df_fenix = pd.read_excel(ruta_fenix)
df_alm = pd.read_excel(ruta_almacen)
df_rel = pd.read_excel(ruta_relacion)

# Normalización
for df in [df_fenix, df_alm, df_rel]:
    df.columns = df.columns.str.lower().str.strip()

df_fenix['pedido'] = df_fenix['pedido'].astype(str).str.strip()
df_alm['pedido'] = df_alm['pedido'].astype(str).str.strip()
df_rel['mano_obra'] = df_rel['mano_obra'].astype(str).str.strip()
df_rel['material_obligatorio'] = df_rel['material_obligatorio'].astype(str).str.strip()

# ------------------------------------------------------------
# VALIDAR MANOS DE OBRA DUPLICADAS
# ------------------------------------------------------------

df_mo_duplicadas = (
    df_alm
    .groupby(
        ['pedido', 'mano_obra']
    )
    .size()
    .reset_index(name='veces')
)

df_mo_duplicadas = df_mo_duplicadas[
    df_mo_duplicadas['veces'] > 1
]

if not df_mo_duplicadas.empty:

    df_mo_duplicadas['estado'] = (
        '🚨 Mano de obra duplicada'
    )
    
# ------------------------------------------------------------
# VALIDACIÓN PRINCIPAL
# ------------------------------------------------------------
resultados = []

for _, fila in df_fenix.iterrows():
    pedido = str(fila['pedido']).strip()
    registros = df_alm[df_alm['pedido'] == pedido]

    
    if registros.empty:
        resultados.append({
            'pedido': pedido,
            'mano_obra': '-',
            'materiales_obligatorios': '-',
            'materiales_entregados': '-',
            'faltantes': '-',
            'estado': '🚨 Pedido no existe en almacén (validar digitación)'
        })
        continue

    mano_obra = registros['mano_obra'].iloc[0]
    entregados = registros['codigo_material'].tolist()

    obligatorios = df_rel[df_rel['mano_obra'] == mano_obra]['material_obligatorio'].tolist()
    faltantes = [m for m in obligatorios if m not in entregados]

    if not obligatorios:
        estado = "⚠️ Mano de obra sin definición"
    elif not faltantes:
        estado = "✅ Pedido completo con materiales"
    else:
        estado = "⚠️ Faltan materiales obligatorios"

    resultados.append({
        'pedido': pedido,
        'mano_obra': mano_obra,
        'materiales_obligatorios': ', '.join(obligatorios) if obligatorios else '-',
        'materiales_entregados': ', '.join(entregados),
        'faltantes': ', '.join(faltantes) if faltantes else '-',
        'estado': estado
    })

# ------------------------------------------------------------
# EXPORTAR RESULTADO
# ------------------------------------------------------------
df_out = pd.DataFrame(resultados)

ruta_salida.parent.mkdir(
    parents=True,
    exist_ok=True
)
print("\n========== MO DUPLICADAS ==========")
print(df_mo_duplicadas.head(20))
print(f"Total duplicadas: {len(df_mo_duplicadas)}")

with pd.ExcelWriter(
    ruta_salida,
    engine="openpyxl"
) as writer:

    df_out.to_excel(
        writer,
        sheet_name="VALIDACION",
        index=False
    )

    df_mo_duplicadas.to_excel(
        writer,
        sheet_name="MO_DUPLICADAS",
        index=False
    )

# ------------------------------------------------------------
# FORMATO VISUAL – TABLA + COLORES ESTILO DASHBOARD (sin cuadricula)
# ------------------------------------------------------------
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook

wb = load_workbook(ruta_salida)

ws = wb["VALIDACION"]



# Crear tabla estructurada limpia
num_filas = ws.max_row
num_cols = ws.max_column
letra_final = get_column_letter(num_cols)

tabla = Table(displayName="VALIDACION_EXPORT", ref=f"A1:{letra_final}{num_filas}")
estilo = TableStyleInfo(
    name="TableStyleMedium2",  # limpio, sin rayas
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=False,
    showColumnStripes=False
)
tabla.tableStyleInfo = estilo
ws.add_table(tabla)

# Quitar bordes visibles
no_border = Border(left=Side(border_style=None),
                   right=Side(border_style=None),
                   top=Side(border_style=None),
                   bottom=Side(border_style=None))

# Aplicar formato a todas las celdas
col_estado = 6  # Columna F
for i in range(2, num_filas + 1):
    celda_estado = ws.cell(row=i, column=col_estado)
    texto = str(celda_estado.value)

    # Alinear a la izquierda (tipo ANS)
    celda_estado.alignment = Alignment(horizontal="left", vertical="center")

    if "✅" in texto:
        celda_estado.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")  # Verde fuerte
        celda_estado.font = Font(color="FFFFFF", bold=True)
    elif "⚠️" in texto:
        celda_estado.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # Amarillo semáforo
        celda_estado.font = Font(color="000000", bold=True)
    elif "🚨" in texto:
        celda_estado.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")  # Rojo fuerte
        celda_estado.font = Font(color="FFFFFF", bold=True)

# Quitar bordes de TODA la hoja y justificar texto a la izquierda
for fila in ws.iter_rows(min_row=2, max_row=num_filas, min_col=1, max_col=num_cols):
    for celda in fila:
        celda.border = no_border
        celda.alignment = Alignment(horizontal="left", vertical="center")

# Ajustar ancho de columnas automáticamente
for col in range(1, num_cols + 1):
    ws.column_dimensions[get_column_letter(col)].auto_size = True

# Desactivar líneas de cuadrícula (vista limpia tipo dashboard)
ws.sheet_view.showGridLines = False

wb.save(ruta_salida)
wb.close()

print("✅ Validación con formato limpio (sin cuadrícula y justificado a la izquierda).")
print("Archivo generado:", ruta_salida)
print(
    f"🚨 Manos de obra duplicadas encontradas: "
    f"{len(df_mo_duplicadas)}"
)