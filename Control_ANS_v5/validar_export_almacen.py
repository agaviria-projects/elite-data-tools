# ============================================================
# VALIDAR EXPORT ALMACÉN
# Nuevo propósito:
# Validación Mano de Obra vs Materiales obligatorios
# ============================================================

from pathlib import Path
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule


def normalizar_mo(valor):
    if pd.isna(valor):
        return None

    import re

    valor = str(valor).upper().strip()

    # Captura A02U, A02, C04U, C04
    match = re.match(r"^([ABC]\d{2})", valor)
    if match:
        return match.group(1)

    # Captura B15R, B17R, etc.
    match = re.match(r"^(B\d{2}R)", valor)
    if match:
        return match.group(1)

    return valor


def normalizar_material(valor):
    if pd.isna(valor):
        return None

    return str(valor).upper().strip()


print("🚀 INICIANDO VALIDACIÓN MANO DE OBRA vs MATERIALES")
print("📍 Archivo ejecutado:", __file__)

# ============================================================
# RUTAS
# ============================================================

BASE_DIR = Path(__file__).resolve().parent
ruta_base_mo = BASE_DIR / "data_master" / "RELACION_MO_MATERIALES.xlsx"

# ============================================================
# CARGA BASE MAESTRA (INSPECCIÓN)
# ============================================================

if not ruta_base_mo.exists():
    raise FileNotFoundError(f"No se encontró la base maestra: {ruta_base_mo}")

df_base = pd.read_excel(ruta_base_mo, dtype=str)

print("📘 Columnas detectadas en la base maestra:")
print(list(df_base.columns))

print("\n📄 Primeras filas:")
print(df_base.head(10))

# ============================================================
# NORMALIZAR BASE MAESTRA MO → LISTA DE MATERIALES
# ============================================================

# Identificar columnas
col_mo = "CÓDIGO ÍTEM MO"

COLS_NO_MATERIAL = set()

cols_materiales = [c for c in df_base.columns if c not in {col_mo, *COLS_NO_MATERIAL}]


# Limpiar y construir lista de materiales por MO
df_base_norm = (
    df_base.assign(
        materiales_obligatorios=lambda d: d[cols_materiales].apply(
            lambda fila: [
                normalizar_material(x)
                for x in fila
                if pd.notna(x) and str(x).strip().upper() != "N/A"
            ],
            axis=1,
        )
    )[[col_mo, "materiales_obligatorios"]]
)

print("\n🧩 Base maestra normalizada:")
print(df_base_norm[df_base_norm[col_mo].str.strip().str.upper().eq("A05")])
print(df_base_norm.head(10))

# ============================================================
# CARGA AUTOMÁTICA DE TXT DE MO vs MATERIALES
# ============================================================

ruta_txt = BASE_DIR / "data_mo_materiales"

if not ruta_txt.exists():
    raise FileNotFoundError("No existe la carpeta data_mo_materiales")

archivos_txt = list(ruta_txt.glob("*.txt"))

if not archivos_txt:
    raise FileNotFoundError("No hay archivos TXT en data_mo_materiales")

print(f"📂 Archivos TXT encontrados: {len(archivos_txt)}")

df_list = []

for archivo in archivos_txt:
    print(f"📄 Leyendo: {archivo.name}")

    df_tmp = pd.read_csv(
        archivo,
        sep="|",            # ⚠️ ajusta si el separador es otro
        dtype=str,
        encoding="latin-1"  # típico en FENIX
    )

    df_tmp["ARCHIVO_ORIGEN"] = archivo.name
    df_list.append(df_tmp)

df_export = pd.concat(df_list, ignore_index=True)

print("\n📦 Columnas detectadas en TXT:")
print(list(df_export.columns))

print("\n📄 Primeras filas consolidadas:")
print(df_export.head(10))

print("\n📦 Columnas detectadas en ALMACEN_EXPORT:")
print(list(df_export.columns))

print("\n📄 Primeras filas ALMACEN_EXPORT:")
print(df_export.head(10))

# ============================================================
# VALIDACIÓN MO vs MATERIALES (LÓGICA PRINCIPAL)
# ============================================================

# ================================
# 1) MANO DE OBRA (solo CON)
# ================================
df_mo = df_export[df_export["tipo"] == "CON"].copy()
df_mo["MO_BASE"] = df_mo["item_cont"].apply(normalizar_mo)

df_mo = df_mo[
    df_mo["pedido"].notna() &
    df_mo["MO_BASE"].notna()
][["pedido", "subz", "MO_BASE"]]

# ================================
# 2) MATERIALES (solo SUM)
# ================================
df_mat = df_export[df_export["tipo"] == "SUM"].copy()
df_mat["MATERIAL"] = df_mat["item_res"].apply(normalizar_material)

df_mat = df_mat[
    df_mat["pedido"].notna() &
    df_mat["MATERIAL"].notna()
][["pedido", "MATERIAL"]]

# df_mat["MATERIAL"] = df_mat["MATERIAL"].str.strip()

df_export_con = df_export[
    (df_export["tipo"] == "CON") &
    (df_export["item_cont"].notna())
].copy()

# 4️⃣ Agrupar materiales entregados por pedido + MO
df_entregados = (
    df_mo
    .merge(df_mat, on="pedido", how="left")
    .groupby(["pedido", "subz", "MO_BASE"])["MATERIAL"]
    .apply(lambda x: sorted(set(x.dropna())))
    .reset_index(name="materiales_entregados")
)

# --- Preparar base maestra como diccionario
mo_to_materiales = dict(
    zip(
        df_base_norm["CÓDIGO ÍTEM MO"],
        df_base_norm["materiales_obligatorios"]
    )
)

resultados = []

for _, row in df_entregados.iterrows():
    pedido = row["pedido"]
    subzona = row["subz"]
    mo = row["MO_BASE"]
    entregados = set(row["materiales_entregados"])

    if mo not in mo_to_materiales:
        resultados.append({
            "pedido": pedido,
            "subzona": row["subz"],
            "mano_obra": mo,
            "estado": "Código de mano de obra no existe en la base maestra",
            "estado_codigo": "NO EXISTEN EN BD",
            "faltantes": "",
            "sobrantes": ", ".join(entregados)
        })
        continue

    obligatorios = set(mo_to_materiales[mo])

    faltantes = sorted(obligatorios - entregados)
    sobrantes = sorted(entregados - obligatorios)

    if not faltantes and not sobrantes:
        estado = "Materiales correctos"
        estado_codigo = "OK"
    elif faltantes and not sobrantes:
        estado = "Faltan materiales obligatorios"
        estado_codigo = "FALTAN"
    elif sobrantes and not faltantes:
        estado = "Materiales sobrantes"
        estado_codigo = "SOBRAN"
    else:
        estado = "Faltan y sobran materiales"
        estado_codigo = "AMBOS"

    resultados.append({
        "pedido": pedido,
        "subzona": subzona,
        "mano_obra": mo,
        "estado": estado,
        "estado_codigo": estado_codigo,
        "faltantes": ", ".join(faltantes),
        "sobrantes": ", ".join(sobrantes)
    })

df_resultado = pd.DataFrame(resultados)

print("\n📊 RESULTADO VALIDACIÓN (vista previa):")
print(df_resultado.head(20))

# ============================================================
# BLOQUE NUEVO (NO TOCA LÓGICA EXISTENTE):
# Reglas especiales "UNO DE", ignorar materiales, y alerta cantidad>1 para MO Cxx
# ============================================================

# --- 1) Reglas "UNO DE" (si aparece cualquiera del grupo, el grupo se considera cumplido)
REGLAS_UNO_DE = {
    # A12: si aparece 200092 o 200093 => OK (no faltante por ese grupo)
    "A12": [{"200092", "200093"}],

    # A05: UNO DE LOS CINCO (incluye 323739)
    "A05": [{"200492", "200410", "200411", "200493", "323739"}],

    # A23: si aparece uno de estos 4 => OK
    "A23": [{"200493", "200411", "200492", "200410"}],

    # A17: si aparece 210954 o 210949 => OK
    "A17": [{"210954", "210949"}],

    # A22: si aparece uno de estos 4 => OK
    "A22": [{"200410", "200493", "200411", "200492"}],

    # A31: si aparece 211618 o 336759 => OK
    "A31": [{"211618", "336759"}],

    # A06: UNO DE LOS CINCO (incluye 323739)
    "A06": [{"200493", "200411", "200492", "200410", "323739"}],

    # A08: UNO DE LOS CUATRO
    "A08": [{"200410", "200493", "200411", "200492"}],
}

# --- 2) Materiales a ignorar (amarillos) si aparecen, para que NO cuenten como sobrantes
MATERIALES_IGNORAR_GLOBAL = {"215887", "219404"}


def _recalcular_estado(obligatorios_set, entregados_set):
    """Replica la lógica de estado sin tocar el loop original."""
    faltantes = sorted(obligatorios_set - entregados_set)
    sobrantes = sorted(entregados_set - obligatorios_set)

    if not faltantes and not sobrantes:
        return ("Materiales correctos", "OK", "", "")
    if faltantes and not sobrantes:
        return ("Faltan materiales obligatorios", "FALTAN", ", ".join(faltantes), "")
    if sobrantes and not faltantes:
        return ("Materiales sobrantes", "SOBRAN", "", ", ".join(sobrantes))

    return ("Faltan y sobran materiales", "AMBOS", ", ".join(faltantes), ", ".join(sobrantes))


def _aplicar_reglas_uno_de(mo, obligatorios_set, entregados_set):
    """
    Si existe regla UNO_DE para la MO:
      - si el entregado cumple algún grupo (intersección), entonces ese grupo queda 'cumplido'
      - por tanto se eliminan como faltantes los elementos faltantes de ese grupo
    """
    grupos = REGLAS_UNO_DE.get(mo, [])
    if not grupos:
        return obligatorios_set  # sin cambios

    obligatorios_corregidos = set(obligatorios_set)

    for grupo in grupos:
        # Si entregados tiene al menos uno del grupo, el grupo se considera cumplido
        if entregados_set.intersection(grupo):
            # Entonces NO exigimos los otros del grupo
            obligatorios_corregidos -= (grupo - entregados_set)

    return obligatorios_corregidos


# --- 3) Mapa rápido de entregados por (pedido, subz, mo)
_entregados_map = {}
for _, r in df_entregados.iterrows():
    k = (r["pedido"], r["subz"], r["MO_BASE"])
    _entregados_map[k] = set(r["materiales_entregados"] or [])


# --- 4) Post-proceso de df_resultado: corrige A05/A12/A23 y aplica ignorados para Cxx (sobrantes)
def _post_procesar_validacion(df_resultado_in):
    df = df_resultado_in.copy()

    for c in ["faltantes", "sobrantes", "estado", "estado_codigo", "mano_obra", "pedido", "subzona"]:
        if c in df.columns:
            df[c] = df[c].fillna("").astype(str)

    mask_okbd = df["estado_codigo"] != "NO EXISTEN EN BD"

    for i, row in df[mask_okbd].iterrows():
        pedido = row["pedido"]
        subz = row["subzona"]
        mo = row["mano_obra"]

        key = (pedido, subz, mo)
        entregados = set(_entregados_map.get(key, set()))

        if mo.startswith("C"):
            entregados = set([x for x in entregados if x not in MATERIALES_IGNORAR_GLOBAL])

        if key not in _entregados_map:
            continue

        obligatorios = set(mo_to_materiales.get(mo, []))
        obligatorios = _aplicar_reglas_uno_de(mo, obligatorios, entregados)

        estado, estado_codigo, faltantes_str, sobrantes_str = _recalcular_estado(obligatorios, entregados)

        # ============================================================
        # 🔴 ALERTA ESPECÍFICA A31 (conflicto + cantidad)
        # ============================================================

        if mo == "A31":

            tiene_211618 = "211618" in entregados
            tiene_336759 = "336759" in entregados

            # 🔴 PRIORIDAD 1: CONFLICTO
            if tiene_211618 and tiene_336759:
                estado = "Revisar: A31 tiene ambos códigos (211618 y 336759)"
                estado_codigo = "REVISAR_A31"
                faltantes_str = ""
                sobrantes_str = ""

            else:
                # 🟠 PRIORIDAD 2: CANTIDAD > 1
                df_sum = df_export[df_export["tipo"] == "SUM"]
                df_sum = df_sum[df_sum["pedido"] == pedido].copy()

                df_sum["MAT"] = df_sum["item_res"].astype(str).str.strip()
                df_sum["CANT_NUM"] = pd.to_numeric(df_sum["cantidad"], errors="coerce").fillna(0)

                df_a31 = df_sum[df_sum["MAT"].isin(["211618", "336759"])]
                mayores = df_a31[df_a31["CANT_NUM"] > 1]

                if not mayores.empty:
                    detalle = ", ".join(
                        [f"{r['MAT']}={int(r['CANT_NUM'])}" for _, r in mayores.iterrows()]
                    )

                    estado = f"Revisar cantidades A31: {detalle}"
                    estado_codigo = "CANTIDAD_A31>1"
                    faltantes_str = ""
                    sobrantes_str = ""

        # 🔥 GUARDAR CAMBIOS (CRÍTICO)
        df.at[i, "estado"] = estado
        df.at[i, "estado_codigo"] = estado_codigo
        df.at[i, "faltantes"] = faltantes_str
        df.at[i, "sobrantes"] = sobrantes_str

    return df


# 🔥 ESTA LÍNEA NO SE BORRA
df_resultado = _post_procesar_validacion(df_resultado)

# ============================================================
# BLOQUE NUEVO: si materiales están en BD pero se consideran "NO APLICAN",
# eliminarlos de faltantes/sobrantes para MO que empiezan con C
# ============================================================

def _quitar_no_aplican_en_resultado(df_in, no_aplican_set):
    df = df_in.copy()
    df["faltantes"] = df["faltantes"].fillna("").astype(str)
    df["sobrantes"] = df["sobrantes"].fillna("").astype(str)

    mask_c = df["mano_obra"].fillna("").astype(str).str.startswith("C")

    def _filtrar_lista_txt(txt):
        items = [x.strip() for x in txt.split(",") if x.strip()]
        items = [x for x in items if x not in no_aplican_set]
        return ", ".join(items)

    df.loc[mask_c, "faltantes"] = df.loc[mask_c, "faltantes"].apply(_filtrar_lista_txt)
    df.loc[mask_c, "sobrantes"] = df.loc[mask_c, "sobrantes"].apply(_filtrar_lista_txt)

    # Recalcular estado/estado_codigo basado en strings (simple)
    for i, r in df[mask_c].iterrows():
        f = r["faltantes"].strip()
        s = r["sobrantes"].strip()

        if not f and not s:
            df.at[i, "estado"] = "Materiales correctos"
            df.at[i, "estado_codigo"] = "OK"
        elif f and not s:
            df.at[i, "estado"] = "Faltan materiales obligatorios"
            df.at[i, "estado_codigo"] = "FALTAN"
        elif s and not f:
            df.at[i, "estado"] = "Materiales sobrantes"
            df.at[i, "estado_codigo"] = "SOBRAN"
        else:
            df.at[i, "estado"] = "Faltan y sobran materiales"
            df.at[i, "estado_codigo"] = "AMBOS"

    return df
# ============================================================
# 🆕 EXCEPCIÓN POR MO:
# Para item C04 se debe EXCLUIR el material 215887A
# (no debe aparecer como faltante ni sobrante SOLO para C04)
# ============================================================

MATERIALES_EXCLUIR_POR_MO = {
    "C04": {"215887A"},
}

def _excluir_materiales_por_mo_en_resultado(df_in, reglas_exclusion):
    df = df_in.copy()

    # Asegurar columnas texto
    for c in ["mano_obra", "faltantes", "sobrantes", "estado", "estado_codigo"]:
        if c in df.columns:
            df[c] = df[c].fillna("").astype(str)

    def _filtrar_txt_por_set(txt, excluir_set):
        items = [x.strip() for x in txt.split(",") if x.strip()]
        items = [x for x in items if x not in excluir_set]
        return ", ".join(items)

    # Aplicar exclusión solo a los MO definidos
    for mo, excluir_set in reglas_exclusion.items():
        mask_mo = df["mano_obra"].str.upper().str.strip().eq(str(mo).upper().strip())

        if "faltantes" in df.columns:
            df.loc[mask_mo, "faltantes"] = df.loc[mask_mo, "faltantes"].apply(
                lambda t: _filtrar_txt_por_set(t, excluir_set)
            )

        if "sobrantes" in df.columns:
            df.loc[mask_mo, "sobrantes"] = df.loc[mask_mo, "sobrantes"].apply(
                lambda t: _filtrar_txt_por_set(t, excluir_set)
            )

        # Recalcular estado basado en strings finales
        for i, r in df[mask_mo].iterrows():
            f = (r.get("faltantes", "") or "").strip()
            s = (r.get("sobrantes", "") or "").strip()

            if not f and not s:
                df.at[i, "estado"] = "Materiales correctos"
                df.at[i, "estado_codigo"] = "OK"
            elif f and not s:
                df.at[i, "estado"] = "Faltan materiales obligatorios"
                df.at[i, "estado_codigo"] = "FALTAN"
            elif s and not f:
                df.at[i, "estado"] = "Materiales sobrantes"
                df.at[i, "estado_codigo"] = "SOBRAN"
            else:
                df.at[i, "estado"] = "Faltan y sobran materiales"
                df.at[i, "estado_codigo"] = "AMBOS"

    return df


df_resultado = _excluir_materiales_por_mo_en_resultado(df_resultado, MATERIALES_EXCLUIR_POR_MO)
df_resultado = _quitar_no_aplican_en_resultado(df_resultado, MATERIALES_IGNORAR_GLOBAL)

# ============================================================
# 🧪 QA: Confirmar exclusión 215887A SOLO para C04
# ============================================================

mask_c04 = df_resultado["mano_obra"].fillna("").astype(str).str.upper().str.strip().eq("C04")

c04_con_215887a_falt = df_resultado.loc[
    mask_c04 & df_resultado["faltantes"].fillna("").astype(str).str.contains(r"\b215887A\b", na=False),
    ["pedido", "subzona", "mano_obra", "estado_codigo", "faltantes"]
]

total_c04 = int(mask_c04.sum())
total_c04_con_215887a = int(len(c04_con_215887a_falt))

print(f"🧪 QA C04 -> Total filas C04: {total_c04}")
print(f"🧪 QA C04 -> Aún con 215887A en FALTANTES: {total_c04_con_215887a}")

if total_c04_con_215887a > 0:
    print("⚠️ Ejemplos donde todavía aparece 215887A como faltante (debería ser 0):")
    print(c04_con_215887a_falt.head(10).to_string(index=False))
else:
    print("✅ OK: se elimina de la BD el material 215887A no es faltante para C04.")


# ============================================================
# ALERTA: Para MO que comiencen con C, si cantidad > 1 en la extracción => marcar alerta
# ============================================================

def _alertas_cantidad_c(df_export_in):
    df_con = df_export_in[df_export_in["tipo"] == "CON"].copy()
    df_con["MO_BASE"] = df_con["item_cont"].apply(normalizar_mo)

    if "cantidad" in df_con.columns:
        df_con["CANT_NUM"] = pd.to_numeric(df_con["cantidad"], errors="coerce").fillna(0)
    else:
        df_con["CANT_NUM"] = 0

    df_con = df_con[df_con["MO_BASE"].astype(str).str.startswith("C")]

    df_alert = df_con[df_con["CANT_NUM"] > 1][["pedido", "subz", "MO_BASE", "CANT_NUM"]].copy()

    if df_alert.empty:
        return pd.DataFrame(columns=["pedido", "subzona", "mano_obra", "ALERTA_CANTIDAD", "DETALLE_CANTIDAD"])

    df_alert["DET"] = df_alert.apply(lambda r: f"{r['MO_BASE']}={int(r['CANT_NUM'])}", axis=1)

    out = (
        df_alert
        .groupby(["pedido", "subz", "MO_BASE"])["DET"]
        .apply(lambda x: ", ".join(sorted(set(x))))
        .reset_index()
        .rename(columns={"subz": "subzona", "MO_BASE": "mano_obra"})
    )

    out["ALERTA_CANTIDAD"] = "CANTIDAD>1"
    out["DETALLE_CANTIDAD"] = out["DET"]
    out = out.drop(columns=["DET"])

    return out[["pedido", "subzona", "mano_obra", "ALERTA_CANTIDAD", "DETALLE_CANTIDAD"]]


df_alertas = _alertas_cantidad_c(df_export_con)

# Merge de alertas (sin romper nada)
if not df_alertas.empty:
    df_resultado = df_resultado.merge(
        df_alertas,
        on=["pedido", "subzona", "mano_obra"],
        how="left"
    )
else:
    df_resultado["ALERTA_CANTIDAD"] = ""
    df_resultado["DETALLE_CANTIDAD"] = ""

df_resultado["ALERTA_CANTIDAD"] = df_resultado["ALERTA_CANTIDAD"].fillna("")
df_resultado["DETALLE_CANTIDAD"] = df_resultado["DETALLE_CANTIDAD"].fillna("")

# ============================================================
# 🆕 ALERTAS ADICIONALES (NO TOCA LÓGICA EXISTENTE)
# 1) Duplicados de Mano de Obra por pedido (ej: mismo pedido con 2 C01)
# 2) Para MO que comiencen con A: si cantidad > 1 => alerta (igual lógica que Cxx)
# ============================================================

def _alerta_duplicados_mo(df_export_in):
    """
    Detecta duplicados de MO_BASE por pedido + subz.
    Aplica para cualquier mano de obra:
    Axx, Bxx, Cxx, B15R, B17R, etc.
    """

    df_con = df_export_in[
        df_export_in["tipo"] == "CON"
    ].copy()

    df_con["MO_BASE"] = (
        df_con["item_cont"]
        .apply(normalizar_mo)
    )

    df_con = df_con[
        df_con["pedido"].notna() &
        df_con["subz"].notna() &
        df_con["MO_BASE"].notna()
    ][["pedido", "subz", "MO_BASE"]]

    df_cnt = (
        df_con
        .groupby(
            ["pedido", "subz", "MO_BASE"]
        )
        .size()
        .reset_index(name="REP")
    )

    df_dup = df_cnt[
        df_cnt["REP"] > 1
    ].copy()

    if df_dup.empty:
        return pd.DataFrame(
            columns=[
                "pedido",
                "subzona",
                "mano_obra",
                "ALERTA_DUPLICADO_MO",
                "DETALLE_DUPLICADO_MO"
            ]
        )

    df_dup["DET"] = df_dup.apply(
        lambda r: f"{r['MO_BASE']} x{int(r['REP'])}",
        axis=1
    )

    out = (
        df_dup
        .groupby(["pedido", "subz"])["DET"]
        .apply(lambda x: ", ".join(sorted(set(x))))
        .reset_index()
        .rename(columns={"subz": "subzona"})
    )

    out["ALERTA_DUPLICADO_MO"] = "MO_DUPLICADA"
    out["DETALLE_DUPLICADO_MO"] = out["DET"]

    out = out.drop(columns=["DET"])

    out["mano_obra"] = ""

    return out[
        [
            "pedido",
            "subzona",
            "mano_obra",
            "ALERTA_DUPLICADO_MO",
            "DETALLE_DUPLICADO_MO"
        ]
    ]


def _alertas_cantidad_prefijo(df_export_in, prefijo="A"):
    """
    Misma lógica de Cxx cantidad>1, pero para prefijo (ej: 'A').
    """
    df_con = df_export_in[df_export_in["tipo"] == "CON"].copy()
    df_con["MO_BASE"] = df_con["item_cont"].apply(normalizar_mo)

    if "cantidad" in df_con.columns:
        df_con["CANT_NUM"] = pd.to_numeric(
            df_con["cantidad"]
            .astype(str)
            .str.replace(",", ".", regex=False)
            .str.strip(),
            errors="coerce"
        ).fillna(0)
    else:
        df_con["CANT_NUM"] = 0
    print("\n========== DEBUG CANTIDADES A ==========")

    print(
        df_con[
            ["pedido", "subz", "MO_BASE", "cantidad", "CANT_NUM"]
        ]
        .query("MO_BASE.str.startswith('A')", engine="python")
        .head(30)
        .to_string(index=False)
    )    
    print(df_con["MO_BASE"].unique())

    df_con = df_con[
        df_con["pedido"].notna() &
        df_con["subz"].notna() &
        df_con["MO_BASE"].notna()
    ][["pedido", "subz", "MO_BASE", "CANT_NUM"]]

    df_con = df_con[df_con["MO_BASE"].astype(str).str.startswith(prefijo)]

    df_alert = df_con[df_con["CANT_NUM"] > 1].copy()
    if df_alert.empty:
        return pd.DataFrame(columns=["pedido", "subzona", "mano_obra", "ALERTA_CANTIDAD_A", "DETALLE_CANTIDAD_A"])

    df_alert["DET"] = df_alert.apply(lambda r: f"{r['MO_BASE']}={int(r['CANT_NUM'])}", axis=1)

    out = (
        df_alert.groupby(["pedido", "subz", "MO_BASE"])["DET"]
        .apply(lambda x: ", ".join(sorted(set(x))))
        .reset_index()
        .rename(columns={"subz": "subzona", "MO_BASE": "mano_obra"})
    )

    out["ALERTA_CANTIDAD_A"] = "CANTIDAD_A>1"
    out["DETALLE_CANTIDAD_A"] = out["DET"]
    out = out.drop(columns=["DET"])

    return out[["pedido", "subzona", "mano_obra", "ALERTA_CANTIDAD_A", "DETALLE_CANTIDAD_A"]]


# --- 1) Ejecutar alerta duplicados (solo C por defecto)
df_dup_mo = _alerta_duplicados_mo(
    df_export
)

# Merge por pedido/subzona (aplica a todas las filas de ese pedido/subzona)
if not df_dup_mo.empty:
    df_resultado = df_resultado.merge(
        df_dup_mo[["pedido", "subzona", "ALERTA_DUPLICADO_MO", "DETALLE_DUPLICADO_MO"]],
        on=["pedido", "subzona"],
        how="left"
    )
else:
    df_resultado["ALERTA_DUPLICADO_MO"] = ""
    df_resultado["DETALLE_DUPLICADO_MO"] = ""

df_resultado["ALERTA_DUPLICADO_MO"] = df_resultado["ALERTA_DUPLICADO_MO"].fillna("")
df_resultado["DETALLE_DUPLICADO_MO"] = df_resultado["DETALLE_DUPLICADO_MO"].fillna("")


# --- 2) Ejecutar alerta cantidad>1 para Axx
df_alertas_a = _alertas_cantidad_prefijo(df_export, prefijo="A")

print("\n========== DEBUG ALERTAS A ==========")
print(df_alertas_a.head(20).to_string(index=False))
print("Total alertas A:", len(df_alertas_a))

# Merge por pedido/subzona/mano_obra (igual que Cxx)
if not df_alertas_a.empty:
    df_resultado = df_resultado.merge(
        df_alertas_a,
        on=["pedido", "subzona", "mano_obra"],
        how="left"
    )
else:
    df_resultado["ALERTA_CANTIDAD_A"] = ""
    df_resultado["DETALLE_CANTIDAD_A"] = ""

df_resultado["ALERTA_CANTIDAD_A"] = df_resultado["ALERTA_CANTIDAD_A"].fillna("")
df_resultado["DETALLE_CANTIDAD_A"] = df_resultado["DETALLE_CANTIDAD_A"].fillna("")


# --- 3) Si hay alerta Axx cantidad>1: estado = novedad cantidades (sin borrar estado_codigo)
mask_alert_a = df_resultado["ALERTA_CANTIDAD_A"].eq("CANTIDAD_A>1")
df_resultado.loc[mask_alert_a, "estado"] = "Presenta novedad en las cantidades (A)"
df_resultado.loc[mask_alert_a, "faltantes"] = ""
df_resultado.loc[mask_alert_a, "sobrantes"] = ""


# ============================================================
# ✅
# - Si hay alerta Cxx cantidad>1:
#     estado = "Presenta novedad en las cantidades"
#     estado_codigo queda igual (OK/FALTAN/SOBRAN/AMBOS)
#     faltantes/sobrantes vacíos
# ============================================================

mask_alert = df_resultado["ALERTA_CANTIDAD"].eq("CANTIDAD>1")
df_resultado.loc[mask_alert, "estado"] = "Presenta novedad en las cantidades"
df_resultado.loc[mask_alert, "faltantes"] = ""
df_resultado.loc[mask_alert, "sobrantes"] = ""


# ============================================================
# DIAGNÓSTICO ALERTAS Cxx cantidad>1 (solo info)
# ============================================================

total_alertas = int((df_resultado["ALERTA_CANTIDAD"] == "CANTIDAD>1").sum())
print(f"🚨 Alertas Cxx por cantidad>1 detectadas: {total_alertas}")

if total_alertas > 0:
    print("📌 Ejemplos de alertas (pedido, subzona, mano_obra, DETALLE_CANTIDAD):")
    print(
        df_resultado.loc[df_resultado["ALERTA_CANTIDAD"] == "CANTIDAD>1",
                         ["pedido", "subzona", "mano_obra", "DETALLE_CANTIDAD"]]
        .head(10)
        .to_string(index=False)
    )


# ============================================================
# LIMPIEZA DE SALIDA (solo presentación)
# - Quita columnas de auditoría que “enredan” el informe
# - Ordena columnas clave
# ============================================================

COLUMNAS_SALIDA = [
    "pedido", "subzona", "mano_obra",
    "estado", "estado_codigo",
    "faltantes",
    "ALERTA_CANTIDAD", "DETALLE_CANTIDAD",
    "ALERTA_CANTIDAD_A", "DETALLE_CANTIDAD_A",
    "ALERTA_DUPLICADO_MO", "DETALLE_DUPLICADO_MO",
    "sobrantes"
]

COLUMNAS_SALIDA = [c for c in COLUMNAS_SALIDA if c in df_resultado.columns]
df_resultado = df_resultado[COLUMNAS_SALIDA].copy()

# Renombrar columnas solo para presentación (minúsculas con sufijo en MAYÚSCULA)
df_resultado = df_resultado.rename(columns={
    "ALERTA_CANTIDAD": "alerta_cantidad_C",
    "DETALLE_CANTIDAD": "detalle_cantidad_C",
    "ALERTA_DUPLICADO_MO": "alerta_duplicado_MO",
    "DETALLE_DUPLICADO_MO": "detalle_duplicado_MO",
    "ALERTA_CANTIDAD_A": "alerta_cantidad_A",
    "DETALLE_CANTIDAD_A": "detalle_cantidad_A",
})
# ============================================================
# HOJA DUPLICADOS MO
# ============================================================

df_con_dup = df_export[
    df_export["tipo"] == "CON"
].copy()

df_con_dup["MO_BASE"] = (
    df_con_dup["item_cont"]
    .apply(normalizar_mo)
)

df_mo_duplicadas = (
    df_con_dup
    .groupby(
        ["pedido", "subz", "MO_BASE"]
    )
    .size()
    .reset_index(name="veces")
)

df_mo_duplicadas = df_mo_duplicadas[
    df_mo_duplicadas["veces"] > 1
]

df_mo_duplicadas = df_mo_duplicadas.rename(
    columns={
        "subz": "subzona",
        "MO_BASE": "mano_obra"
    }
)

print(
    f"🚨 Manos de obra duplicadas encontradas: "
    f"{len(df_mo_duplicadas)}"
)

# ============================================================
# HOJA NUEVA: MATERIALES_CANTIDAD
# Detecta materiales SUM con cantidad > 1
# ============================================================

df_material_cantidad = df_export[
    df_export["tipo"] == "SUM"
].copy()

df_material_cantidad["material"] = (
    df_material_cantidad["item_res"]
    .astype(str)
    .str.strip()
)

df_material_cantidad["cantidad"] = pd.to_numeric(
    df_material_cantidad["cantidad"]
    .astype(str)
    .str.replace(",", ".", regex=False)
    .str.strip(),
    errors="coerce"
).fillna(0)

df_material_cantidad = df_material_cantidad[
    df_material_cantidad["cantidad"] > 1
][
    ["pedido", "subz", "material", "cantidad"]
].copy()

df_material_cantidad = df_material_cantidad.rename(
    columns={
        "subz": "subzona"
    }
)

df_material_cantidad["alerta"] = "MATERIAL_CANTIDAD>1"

df_material_cantidad["detalle"] = (
    df_material_cantidad["material"].astype(str)
    + "="
    + df_material_cantidad["cantidad"].astype(str)
)

print(
    f"🚨 Materiales con cantidad > 1 encontrados: "
    f"{len(df_material_cantidad)}"
)
# ============================================================
# EXPORTAR RESULTADO
# ============================================================

ruta_salida = BASE_DIR / "outputs"
ruta_salida.mkdir(exist_ok=True)

archivo = ruta_salida / f"validacion_mo_materiales_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
with pd.ExcelWriter(
    archivo,
    engine="openpyxl"
) as writer:

    df_resultado.to_excel(
        writer,
        sheet_name="VALIDACION",
        index=False
    )

    df_mo_duplicadas.to_excel(
        writer,
        sheet_name="MO_DUPLICADAS",
        index=False
    )
    df_material_cantidad.to_excel(
        writer,
        sheet_name="ALERTA_CANTIDADES",
        index=False
    )
# ============================================================
# FORMATO PROFESIONAL DEL EXCEL (NO TOCA LÓGICA)
# ============================================================

wb = load_workbook(archivo)

if "MO_DUPLICADAS" in wb.sheetnames:
    wb["MO_DUPLICADAS"].sheet_properties.tabColor = "C00000"

if "ALERTA_CANTIDADES" in wb.sheetnames:
    wb["ALERTA_CANTIDADES"].sheet_properties.tabColor = "FFC000"

ws = wb["VALIDACION"]

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

header_fill = PatternFill("solid", fgColor="1E8449")
header_font = Font(color="FFFFFF", bold=True)  # 👈 más pro
header_align = Alignment(horizontal="center", vertical="center")

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align

headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

# --- Pintar en rojo detalle_cantidad_C cuando haya alerta_cantidad_C
col_alerta_c = headers.get("alerta_cantidad_C")
col_detalle_c = headers.get("detalle_cantidad_C")

if col_alerta_c and col_detalle_c:
    fill_rojo = PatternFill("solid", fgColor="FF0000")
    font_blanco = Font(color="FFFFFF", bold=True)

    for r in range(2, ws.max_row + 1):
        v_alerta = ws.cell(row=r, column=col_alerta_c).value
        if str(v_alerta).strip().upper() == "CANTIDAD>1":
            c = ws.cell(row=r, column=col_detalle_c)
            c.fill = fill_rojo
            c.font = font_blanco

# --- Pintar en rojo detalle_duplicado_MO cuando haya MO_DUPLICADA
col_alerta_dup = headers.get("alerta_duplicado_MO")
col_detalle_dup = headers.get("detalle_duplicado_MO")

if col_alerta_dup and col_detalle_dup:
    fill_rojo = PatternFill("solid", fgColor="FF0000")
    font_blanco = Font(color="FFFFFF", bold=True)

    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=col_alerta_dup).value
        if str(v).strip().upper() == "MO_DUPLICADA":
            c = ws.cell(row=r, column=col_detalle_dup)
            c.fill = fill_rojo
            c.font = font_blanco

# --- Pintar en rojo detalle_cantidad_A cuando haya CANTIDAD_A>1
col_alerta_a = headers.get("alerta_cantidad_A")
col_detalle_a = headers.get("detalle_cantidad_A")

if col_alerta_a and col_detalle_a:
    fill_rojo = PatternFill("solid", fgColor="FF0000")
    font_blanco = Font(color="FFFFFF", bold=True)

    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=col_alerta_a).value
        if str(v).strip().upper() == "CANTIDAD_A>1":
            c = ws.cell(row=r, column=col_detalle_a)
            c.fill = fill_rojo
            c.font = font_blanco

# --- Pintar en rojo TODA LA FILA cuando haya CONFLICTO_A31
col_estado_codigo = headers.get("estado_codigo")

if col_estado_codigo:
    fill_rojo = PatternFill("solid", fgColor="FF0000")
    font_blanco = Font(color="FFFFFF", bold=True)

    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=col_estado_codigo).value
        if str(v).strip().upper() == "REVISAR_A31":
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill_rojo
                cell.font = font_blanco

for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)

    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))

    ws.column_dimensions[col_letter].width = min(max_length + 3, 45)

for row in ws.iter_rows(min_row=2):
    row[0].alignment = Alignment(horizontal="center")
    row[1].alignment = Alignment(horizontal="center")
    row[2].alignment = Alignment(horizontal="left")
    row[3].alignment = Alignment(horizontal="left")
    row[4].alignment = Alignment(wrap_text=True)
    row[5].alignment = Alignment(wrap_text=True)
# ============================================================
# FORMATO HOJA MO_DUPLICADAS
# ============================================================

if "MO_DUPLICADAS" in wb.sheetnames:

    ws_dup = wb["MO_DUPLICADAS"]

    # Congelar encabezado
    ws_dup.freeze_panes = "A2"

    # Filtro
    ws_dup.auto_filter.ref = ws_dup.dimensions

    # Encabezado rojo
    fill_rojo = PatternFill(
        "solid",
        fgColor="C00000"
    )

    font_blanco = Font(
        color="FFFFFF",
        bold=True
    )

    align_center = Alignment(
        horizontal="center",
        vertical="center"
    )

    for cell in ws_dup[1]:
        cell.fill = fill_rojo
        cell.font = font_blanco
        cell.alignment = align_center

    # Pintar filas de alerta
    fill_alerta = PatternFill(
        "solid",
        fgColor="FDE9E7"
    )

    for fila in ws_dup.iter_rows(
        min_row=2,
        max_row=ws_dup.max_row,
        min_col=1,
        max_col=ws_dup.max_column
    ):
        for celda in fila:
            celda.fill = fill_alerta

    # Ajustar ancho columnas
    for col in ws_dup.columns:

        max_length = 0

        letra = get_column_letter(
            col[0].column
        )

        for cell in col:
            if cell.value:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        ws_dup.column_dimensions[
            letra
        ].width = min(
            max_length + 5,
            40
        )
# ============================================================
# FORMATO HOJA MATERIAL_CANTIDAD
# ============================================================

if "ALERTA_CANTIDADES" in wb.sheetnames:

    ws_mat = wb["ALERTA_CANTIDADES"]

    ws_mat.freeze_panes = "A2"
    ws_mat.auto_filter.ref = ws_mat.dimensions

    fill_naranja = PatternFill(
        "solid",
        fgColor="FFC000"
    )

    font_negro = Font(
        color="000000",
        bold=True
    )

    align_center = Alignment(
        horizontal="center",
        vertical="center"
    )

    for cell in ws_mat[1]:
        cell.fill = fill_naranja
        cell.font = font_negro
        cell.alignment = align_center

    fill_alerta_mat = PatternFill(
        "solid",
        fgColor="FFF2CC"
    )

    for fila in ws_mat.iter_rows(
        min_row=2,
        max_row=ws_mat.max_row,
        min_col=1,
        max_col=ws_mat.max_column
    ):
        for celda in fila:
            celda.fill = fill_alerta_mat

    for col in ws_mat.columns:

        max_length = 0
        letra = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        ws_mat.column_dimensions[letra].width = min(
            max_length + 5,
            40
        )
        # ==========================
        # BARRAS DE DATOS
        # ==========================

        encabezados = {
            cell.value: idx + 1
            for idx, cell in enumerate(ws_mat[1])
        }

        col_cantidad = encabezados.get("cantidad")

        if col_cantidad:

            letra = get_column_letter(col_cantidad)

            regla = DataBarRule(
                start_type="min",
                end_type="max",
                color="FF6B6B",
                showValue=True
            )

            ws_mat.conditional_formatting.add(
                f"{letra}2:{letra}{ws_mat.max_row}",
                regla
            )

wb.save(archivo)

print(f"📁 Archivo generado: {archivo}")