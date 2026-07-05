import pandas as pd
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


# =====================================================
# RUTAS
# =====================================================

BASE = Path(__file__).resolve().parent

ENTRADA = BASE / "entrada"
SALIDA = BASE / "salida"

ARCHIVO_INTERVENTORIA = ENTRADA / "ACTAS_UNIFICADAS_INTERVENTORIA.xlsx"
ARCHIVO_DRACO = ENTRADA / "DRACO_UNIFICADO.xlsx"

ARCHIVO_SALIDA = SALIDA / "CONCILIACION_DRACO_INTERVENTORIA.xlsx"

# =====================================================
# VALIDAR ARCHIVOS
# =====================================================

if not ARCHIVO_INTERVENTORIA.exists():
    raise FileNotFoundError(
        f"No existe:\n{ARCHIVO_INTERVENTORIA}"
    )

if not ARCHIVO_DRACO.exists():
    raise FileNotFoundError(
        f"No existe:\n{ARCHIVO_DRACO}"
    )

print("✅ Archivos encontrados")

# =====================================================
# LEER ARCHIVOS
# =====================================================

print("📖 Leyendo Interventoría...")
df_int = pd.read_excel(
    ARCHIVO_INTERVENTORIA,
    sheet_name="ACTAS_UNIFICADAS",
    dtype=str
)

df_int.columns = (
    df_int.columns
    .astype(str)
    .str.strip()
)

print("\n📖 Leyendo DRACO...")

df_draco = pd.read_excel(
    ARCHIVO_DRACO,
    dtype=str
)

df_draco.columns = (
    df_draco.columns
    .astype(str)
    .str.strip()
)

# =====================================================
# INTERVENTORIA
# =====================================================

print("🔄 Procesando Interventoría...")

df_int = df_int.rename(columns={
    "pedido": "pedido",
    "acta": "acta_interventoria",
    "suminis": "material",
    "cantidad": "cantidad_interventoria"
})

df_int["pedido"] = (
    df_int["pedido"]
    .astype(str)
    .str.strip()
)

# =====================================================
# FILTRO CONTRATO
# =====================================================

registros_antes = len(df_int)

df_int["contrato"] = (
    df_int["contrato"]
    .astype(str)
    .str.strip()
)

df_int = df_int[
    df_int["contrato"] == "CW352017"
].copy()

print()
print("=" * 50)
print("FILTRO CONTRATO")
print("=" * 50)
print("Antes:", registros_antes)
print("Después:", len(df_int))

# =====================================================
# NORMALIZAR MATERIAL
# =====================================================

df_int["material"] = (
    df_int["material"]
    .astype(str)
    .str.upper()
    .str.strip()
)

# Quitar todo lo que no sea número

df_int["material"] = (
    df_int["material"]
    .str.extract(r"^(\d{6})", expand=False)
)


# Convertir a texto

df_int["material"] = (
    df_int["material"]
    .astype(str)
    .str.strip()
)

# Conservar únicamente materiales de 6 dígitos
# que empiecen por 2

df_int = df_int[
    df_int["material"].str.fullmatch(r"\d{6}", na=False)
].copy()

df_int["cantidad_interventoria"] = pd.to_numeric(
    df_int["cantidad_interventoria"],
    errors="coerce"
).fillna(0)

print()
print("=" * 50)
print("INTERVENTORIA")
print("=" * 50)
print("Registros:", len(df_int))
print("Pedidos:", df_int["pedido"].nunique())
print("Materiales:", df_int["material"].nunique())

print()
print("MUESTRA INTERVENTORIA")

df_int = (
    df_int
    .groupby(
        [
            "pedido",
            "material",
            "acta_interventoria"
        ],
        as_index=False
    )
    ["cantidad_interventoria"]
    .sum()
)
df_int["cantidad_interventoria"] = pd.to_numeric(
    df_int["cantidad_interventoria"],
    errors="coerce"
).fillna(0)

print(f"✅ Interventoría agrupada: {len(df_int):,}")

# =====================================================
# DRACO
# =====================================================

print("🔄 Procesando DRACO...")

df_draco = df_draco.rename(columns={
    "Pedido": "pedido",
    "Acta": "acta_draco",
    "Código": "material",
    "Técnico": "tecnico",
    "Cantidad": "cantidad_draco"
})

df_draco["pedido"] = (
    df_draco["pedido"]
    .astype(str)
    .str.strip()
)

df_draco["material"] = (
    df_draco["material"]
    .astype(str)
    .str.strip()
)

df_draco["material"] = (
    df_draco["material"]
    .str.extract(r"(\d{6})", expand=False)
)

df_draco = df_draco[
    df_draco["material"].str.fullmatch(r"\d{6}", na=False)
].copy()

df_draco["cantidad_draco"] = pd.to_numeric(
    df_draco["cantidad_draco"],
    errors="coerce"
).fillna(0)

print()
print("=" * 50)
print("DRACO")
print("=" * 50)
print("Registros:", len(df_draco))
print("Pedidos:", df_draco["pedido"].nunique())
print("Materiales:", df_draco["material"].nunique())

df_draco = (
    df_draco
    .groupby(
        [
            "pedido",
            "material"
        ],
        as_index=False
    )
    .agg(
        {
            "cantidad_draco": "sum",
            "acta_draco": "first",
            "tecnico": "first"
        }
    )
)

print(f"✅ DRACO agrupado: {len(df_draco):,}")

print()
print("=" * 50)
print("MUESTRA INTERVENTORIA")
print("=" * 50)
print(df_int.head())

print()
print("=" * 50)
print("MUESTRA DRACO")
print("=" * 50)
print(df_draco.head())
# =====================================================
# CRUCE
# =====================================================

print("🔗 Cruzando información...")

df_final = df_int.merge(
    df_draco,
    on=[
        "pedido",
        "material"
    ],
    how="left"
)

df_final["cantidad_draco"] = pd.to_numeric(
    df_final["cantidad_draco"],
    errors="coerce"
).fillna(0)

df_final["diferencia"] = (
    df_final["cantidad_interventoria"]
    - df_final["cantidad_draco"]
)

# =====================================================
# ESTADOS
# =====================================================

def clasificar(row):

    cant_int = row["cantidad_interventoria"]
    cant_dra = row["cantidad_draco"]

    if cant_dra == 0:
        return "🚨 NO REPORTADO EN DRACO"

    if round(cant_int, 2) == round(cant_dra, 2):
        return "✅ OK"

    if cant_dra > cant_int:
        return "⚠ MATERIAL ADICIONAL EN DRACO"

    return "⚠ DIFERENCIA CANTIDAD"


df_final["estado"] = df_final.apply(
    clasificar,
    axis=1
)
print()
print("=" * 50)
print("RESUMEN ESTADOS")
print("=" * 50)
print(df_final["estado"].value_counts())
# =====================================================
# HOJAS AUXILIARES
# =====================================================

df_no_reportados = df_final[
    df_final["estado"]
    == "🚨 NO REPORTADO EN DRACO"
].copy()

df_diferencias = df_final[
    df_final["estado"].isin(
        [
            "⚠ DIFERENCIA CANTIDAD",
            "⚠ MATERIAL ADICIONAL EN DRACO"
        ]
    )
].copy()

# =====================================================
# RESUMEN
# =====================================================

resumen = pd.DataFrame(
    {
        "Metrica": [
            "Registros Interventoria",
            "Registros Draco",
            "Coincidencias",
            "Diferencias",
            "No Reportados"
        ],
        "Valor": [
            len(df_int),
            len(df_draco),
            (df_final["estado"] == "✅ OK").sum(),
            len(df_diferencias),
            len(df_no_reportados)
        ]
    }
)
SALIDA.mkdir(
    parents=True,
    exist_ok=True
)
df_final = df_final.sort_values(
    by=["estado", "pedido", "material"]
)

df_final = df_final[
    [
        "pedido",
        "material",
        "acta_interventoria",
        "cantidad_interventoria",
        "cantidad_draco",
        "acta_draco",
        "tecnico",
        "diferencia",
        "estado"
    ]
]
# =====================================================
# EXPORTAR
# =====================================================

print("💾 Generando Excel...")

with pd.ExcelWriter(
    ARCHIVO_SALIDA,
    engine="openpyxl"
) as writer:

    df_final.to_excel(
        writer,
        sheet_name="CONCILIACION",
        index=False
    )
    ws = writer.sheets["CONCILIACION"]

    # Insertar fila superior
    ws.insert_rows(1)

    # Agrupar columnas
    ws.merge_cells("A1:D1")
    ws.merge_cells("E1:G1")
    ws.merge_cells("H1:I1")

    ws["A1"] = "INTERVENTORÍA"
    ws["E1"] = "DRACO"
    ws["H1"] = "CONCILIACIÓN"


    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20

    azul = PatternFill(
    fill_type="solid",
    fgColor="1F4E78"
    )

    naranja = PatternFill(
        fill_type="solid",
        fgColor="C55A11"
    )

    verde = PatternFill(
        fill_type="solid",
        fgColor="548235"
    )

    for celda, color in [
        ("A1", azul),
        ("E1", naranja),
        ("H1", verde),
    ]:

        ws[celda].fill = color

        ws[celda].font = Font(
            bold=True,
            color="FFFFFF",
            size=12
        )

        ws[celda].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
    for celda in ws[2]:

        celda.font = Font(bold=True)

        celda.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # Congelar encabezados
    ws.freeze_panes = "A3"  

    # Crear tabla de Excel

    ultima_fila = ws.max_row
    ultima_columna = get_column_letter(ws.max_column)

    tabla = Table(
        displayName="TablaConciliacion",
        ref=f"A2:{ultima_columna}{ultima_fila}"
    )

    estilo = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )

    tabla.tableStyleInfo = estilo

    ws.add_table(tabla)

    # =====================================
    # ANCHO DE COLUMNAS
    # =====================================

    ws.column_dimensions["A"].width = 14   # Pedido
    ws.column_dimensions["B"].width = 12   # Material
    ws.column_dimensions["C"].width = 16   # Acta Interventoría
    ws.column_dimensions["D"].width = 18   # Cantidad Interventoría

    ws.column_dimensions["E"].width = 16   # Cantidad DRACO
    ws.column_dimensions["F"].width = 14   # Acta DRACO
    ws.column_dimensions["G"].width = 28   # Técnico

    ws.column_dimensions["H"].width = 14   # Diferencia
    ws.column_dimensions["I"].width = 34   # Estado


  

    # =====================================
    # ALINEACIÓN
    # =====================================

    for fila in ws.iter_rows(min_row=3):

        fila[0].alignment = Alignment(horizontal="center")   # Pedido
        fila[1].alignment = Alignment(horizontal="center")   # Material
        fila[2].alignment = Alignment(horizontal="center")   # Acta
        fila[3].alignment = Alignment(horizontal="center")   # Cantidad Int
        fila[4].alignment = Alignment(horizontal="center")   # Cantidad Draco
        fila[5].alignment = Alignment(horizontal="center")   # Acta Draco
        fila[7].alignment = Alignment(horizontal="center")   # Diferencia

    # =====================================
    # EXPORTAR OTRAS HOJAS
    # =====================================

    df_no_reportados.to_excel(
        writer,
        sheet_name="NO_REPORTADOS",
        index=False
    )

    df_diferencias.to_excel(
        writer,
        sheet_name="DIFERENCIAS",
        index=False
    )

    resumen.to_excel(
        writer,
        sheet_name="RESUMEN",
        index=False
    )
print("ENTRADA:", ENTRADA)
print("INTERVENTORIA:", ARCHIVO_INTERVENTORIA)
print("DRACO:", ARCHIVO_DRACO)
print("=================================")
print("✅ PROCESO FINALIZADO")
print(f"📁 Archivo: {ARCHIVO_SALIDA}")
print("=================================")